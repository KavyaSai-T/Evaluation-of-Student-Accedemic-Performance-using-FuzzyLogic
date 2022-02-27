from django.shortcuts import render, redirect,get_object_or_404
from .models import *
from .forms import *
from django.contrib import messages
from django.http import JsonResponse
from django.contrib.sessions.models import Session
from django.db import connection
from django.core import serializers
from django.http import HttpResponse
from django.db.models import Count
from django.db.models import Sum
from datetime import datetime
from django.http import HttpResponseRedirect
from django.conf.urls import url
import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import itertools 
from django.db.models import Q   
from . performance import *
import random 
import csv

def performance_analysis(request):
	a = request.GET.get('ex')
	b = request.GET.get('internal')
	c = request.GET.get('percentage')
	result=compute_fuzzy(a,b,c)
	return render(request,'calculate_performance.html',{'credit':result})

def cal(request):
    if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
        a1 = request.GET.get('ex')
        b1 = request.GET.get('internal')
        c1 = request.GET.get('percentage')
        total = float(a1)+float(b1)+float(c1)
        avg = int(total/3)
        if (avg<=100 and avg>=90):
            analysis = Excellent(request)
        elif (avg<=89 and avg>=70):
            analysis = Good(request)
        elif (avg<=69 and avg>=50):
            analysis = Average(request)
        elif(avg<=49 and avg>=35):
            analysis = Poor(request)
        else:
            analysis = VeryPoor(request)
        return render(request, 'calculate_performance.html', {'total':avg,'analysis':analysis})
    else:
        return redirect('adminlogin')

def error_404_view(request, exception):
    data = {}
    return render(request,'error_404.html', data)
    
def pagination_cnt(request):
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    sql = '''SELECT app_schoolsetting.pagination_count from app_schoolsetting WHERE 
    app_schoolsetting.school_id_id='%d' ''' % (school_id)
    post = cursor.execute(sql)
    row = cursor.fetchone()
    count = row[0]
    return count
def export_subject_to_xlsx(request):
    if request.method == 'GET':
        class_id = request.GET.get('cls_id')
        section_id = request.GET.get('sec_id')
        count = request.GET.get('count_val')
        staff_id = request.session['user_id']
        a = int(class_id)
        b = int(section_id)
        cnt = int(count)
        subject_queryset = AssignSubjectTeacher.objects.filter(class_id=a,section_id=b,staff_id=staff_id).order_by('subject_id')
        cursor = connection.cursor()
        post = '''SELECT  school_studentdetail.register_number,school_studentdetail.student_name from school_studentsection INNER JOIN school_studentdetail
        ON school_studentsection.student_id_id=school_studentdetail.id where 
        school_studentsection.class_id_id='%d' AND school_studentsection.section_id_id='%d' order by school_studentsection.student_id_id ''' % (a,b)
        query = cursor.execute(post)
        row = cursor.fetchall()
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-marksheet.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Mark Sheet'

        # Define the titles for columns
        columns = [
            'Student Register Number',
           
        ]

        name = ["0", "Student Name"]

        num = ["0","S.No"]

        for a in range(1,cnt+1):
            num.append(a)

        for i in row:
            columns.append(i[0])
            name.append(i[1])
            
        for i in range(1,cnt+2):
            worksheet.cell(row=i, column=1).value = num[i]

        for i in range(1,cnt+2):
            worksheet.cell(row=i, column=2).value = name[i]

        row_num = 3

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=col_num, column=row_num)
            cell.value = column_title

        # Iterate through all movies
        for sub in subject_queryset:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                sub.subject_id.subject_name,
                
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=col_num, column=row_num)
                cell.value = cell_value 

        workbook.save(response)

        return response  
def calculate_performance(request):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		a = request.GET.get('ex')
		b = request.GET.get('internal')
		c = request.GET.get('percentage')
		total = float(a)+float(b)+float(c)
		avg = int(total/3)
		if (avg<=100 and avg>=90):
			credit = "Excellent"
			analysis = Excellent(request)
		elif (avg<=89 and avg>=70):
			credit = "Good"
			analysis = Good(request)
		elif (avg<=69 and avg>=50):
			credit = "Average"
			analysis = Average(request)
		elif (avg<=49 and avg>=35):
			credit = "Poor"
			analysis = Poor(request)
		else:
			credit = "Very Poor"
			analysis = VeryPoor(request)
		return render(request, 'calculate_performance.html', {'total':avg,'credit':credit,'analysis':analysis})
	else:
		return redirect('adminlogin')    
def adminlogin(request):
    if request.session.has_key('username'):
        return redirect("dashboard")
    else:
        if request.method == 'POST':
            username = request.POST['username']
            password =  request.POST['password']
            post = SchoolDetail.objects.filter(username=username,password=password)
            if post:
                username = request.POST['username']
                request.session['username'] = username
                a = request.session['username']
                sess = SchoolDetail.objects.only('id').get(username=a).id
                request.session['schoolname']=sess
                return redirect("dashboard")
            else:
                messages.success(request, 'Invalid Username or Password')
    return render(request, 'index.html', {})
def teacher_login(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        return redirect("teacher_dashboard")
    else:
        if request.method == 'POST':
            username = request.POST['username']
            password =  request.POST['password']
            post = StaffDetail.objects.filter(username=username,password=password)
            if post:
                username = request.POST['username']
                request.session['teacher'] = username
                a = request.session['teacher']
                user_id = StaffDetail.objects.only('staff_id').get(username=a).staff_id
                request.session['user_id']=user_id
                sess = StaffDetail.objects.only('school_id_id').get(staff_id=user_id).school_id_id
                request.session['schoolname']=sess
                return redirect("teacher_dashboard")
            else:
                messages.success(request, 'Invalid Username or Password')
    return render(request, 'teacher_login.html', {})
def student_login(request):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
         return redirect("student_dashboard")
    else:
        if request.method == 'POST':
            username = request.POST['username']
            password =  request.POST['password']
            post = StudentLogin.objects.filter(username=username,password=password)
            if post:
                username = request.POST['username']
                request.session['student'] = username
                a=request.session['student']
                stud_id = StudentLogin.objects.only('login_id').get(username=a).login_id
                request.session['student_id']=stud_id
                stud_reg_no = StudentLogin.objects.only('reg_no').get(username=a).reg_no
                request.session['register_number']=stud_reg_no
                return redirect("student_dashboard")
            else:
                messages.success(request, 'Invalid Username or Password')
    return render(request, 'student_login.html', {})
def dashboard(request):
	if request.session.has_key('username'):
		posts = request.session['username']
		query = SchoolDetail.objects.filter(username=posts)
		return render(request, 'dashboard.html', {"query":query})
	else:
		return render(request, 'index.html',{})
def teacher_dashboard(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        return render(request, 'teacher_dashboard.html',{})
    else:
        return render(request, 'teacher_login.html',{})
def student_dashboard(request):
    if request.session.has_key('student'):
        return render(request, 'student_dashboard.html',{})
    else:
        return render(request, 'student_login.html',{})
def logout(request):
    try:
        del request.session['username']
        del request.session['academic_year']
    except:
     pass
    return render(request, 'index.html', {})

def teacher_logout(request):
    try:
        del request.session['teacher']
        del request.session['academic_year']
    except:
     pass
    return render(request, 'teacher_login.html', {})

def student_logout(request):
    try:
        del request.session['student']
        del request.session['academic_year']
    except:
     pass
    return render(request, 'student_login.html', {})
def add_teacher(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        if request.method == "POST":
            form = StaffForm(request.POST, request.FILES)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                return redirect('teachers')
        else:
            form = StaffForm()
        
        return render(request, 'add_teacher.html', {'forms':form})
    else:
        return redirect("adminlogin")
def teachers(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        teacher = StaffDetail.objects.filter(school_id_id=school_id).order_by('-staff_id')
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(teacher,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
        return render(request, 'teachers.html', {'teacher':users})
    else:
        return redirect("adminlogin")
def add_student(request):
	if request.session.has_key('username') and request.session.has_key('academic_year'):
		if request.method == "POST":
			school_id = request.POST.get('school_id')
			ids = SchoolDetail.objects.get(id=int(school_id))
			student_name = request.POST.get('student_name')
			register_number = request.POST.get('register_number')
			admission_number = request.POST.get('admission_number')
			date_of_admission = request.POST.get('date_of_admission')
			date_of_birth = request.POST.get('date_of_birth')
			adhar_number = request.POST.get('adhar_number')
			father_name = request.POST.get('father_name')
			mother_name = request.POST.get('mother_name')
			mother_name = request.POST.get('mother_name')
			emergency_number = request.POST.get('emergency_number')
			father_occupation = request.POST.get('father_occupation')
			mother_occupation = request.POST.get('mother_occupation')
			present_address = request.POST.get('present_address')
			permanent_address = request.POST.get('permanent_address')
			religion = request.POST.get('religion')
			caste = request.POST.get('caste')
			blood_group = request.POST.get('blood_group')
			Photo = request.FILES['Photo']
			no_of_backlogs = request.POST.get('no_of_backlogs')
			crt = StudentDetail.objects.create(school_id=ids,student_name=student_name,register_number=register_number,admission_number=admission_number,
			date_of_admission=date_of_admission,date_of_birth=date_of_birth,adhar_number=adhar_number,father_name=father_name,mother_name=mother_name,
			emergency_number=emergency_number,father_occupation=father_occupation,mother_occupation=mother_occupation,
			present_address=present_address,permanent_address=permanent_address,no_of_backlogs=int(no_of_backlogs),religion=religion,caste=caste,blood_group=blood_group,Photo=Photo)
			if crt:
				return redirect('students')

		return render(request, 'add_students.html', {})
	else:
		return redirect("adminlogin")
def studentedit(request,pk):
	if request.session.has_key('username') and request.session.has_key('academic_year'):
		detail = StudentDetail.objects.filter(id=pk)
		if request.method == "POST":
			school_id = request.POST.get('school_id')
			ids = SchoolDetail.objects.get(id=int(school_id))
			student_name = request.POST.get('student_name')
			register_number = request.POST.get('register_number')
			admission_number = request.POST.get('admission_number')
			adhar_number = request.POST.get('adhar_number')
			father_name = request.POST.get('father_name')
			mother_name = request.POST.get('mother_name')
			mother_name = request.POST.get('mother_name')
			emergency_number = request.POST.get('emergency_number')
			father_occupation = request.POST.get('father_occupation')
			mother_occupation = request.POST.get('mother_occupation')
			present_address = request.POST.get('present_address')
			permanent_address = request.POST.get('permanent_address')
			religion = request.POST.get('religion')
			caste = request.POST.get('caste')
			blood_group = request.POST.get('blood_group')
			no_of_backlogs = request.POST.get('no_of_backlogs')
			crt = StudentDetail.objects.filter(id=pk).update(student_name=student_name,register_number=register_number,admission_number=admission_number,
			adhar_number=adhar_number,father_name=father_name,mother_name=mother_name,
			emergency_number=emergency_number,father_occupation=father_occupation,mother_occupation=mother_occupation,
			present_address=present_address,permanent_address=permanent_address,no_of_backlogs=int(no_of_backlogs),religion=religion,caste=caste,blood_group=blood_group)
			if crt:
				return redirect('students')

		return render(request, 'studentedit.html', {'detail':detail})
	else:
		return redirect("adminlogin")
def edit_teacher(request,pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        post = get_object_or_404(StaffDetail, pk=pk)
        if request.method == "POST":
            form = StaffEditForm(request.POST,instance=post)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                return redirect('teachers')
        else:
            form = StaffEditForm(instance=post)
        return render(request, 'edit_teacher.html', {'form': form})
    else:
        return redirect("adminlogin")
def teacher_delete(request, pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        student = StaffDetail.objects.get(staff_id=pk)
        student.delete()
        return redirect("teachers")
    else:
        return redirect("adminlogin")
def students(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        ids = StudentDetail.objects.latest('id')
        school_id = request.session['schoolname']
        student = StudentDetail.objects.filter(school_id_id=school_id).order_by('-id')
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(student,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
        return render(request, 'students.html', {'student':users,'ids':ids})
    else:
        return redirect("adminlogin")
def student_edit(request,pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        post = get_object_or_404(StudentDetail, pk=pk)
        if request.method == "POST":
            form = StudentEditForm(request.POST, request.FILES, instance=post)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                return redirect("students")
        else:
            form = StudentEditForm(instance=post)
        return render(request, 'student_edit.html', {'form': form})
    else:
        return redirect("adminlogin")
def student_delete(request, pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        student = StudentDetail.objects.get(id=pk)
        student.delete()
        return redirect("students")
    else:
        return redirect("adminlogin")

def add_class_section(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        existing_class = ""
        sec_create = ""
        class_id = dict()
        school_id = request.session['schoolname']
        Scl_id = SchoolDetail.objects.get(id=int(school_id))
        academic_year = request.session['academic_year']
        cursor = connection.cursor()
        cls_new = ''' SELECT app_class.class_id,app_class.class_name,app_section.section_id,app_section.section_name
        from app_class LEFT JOIN app_section ON app_class.class_id=app_section.class_id_id where
        app_class.academic_year='%s' AND app_class.school_id_id='%d' order by app_section.section_id DESC''' % (academic_year,school_id)
        post = cursor.execute(cls_new)
        cls_sec_name = cursor.fetchall()

        cls_name = Section.objects.filter(class_id__in=Class.objects.filter(academic_year=academic_year,school_id=school_id)).order_by('-section_id')
        cls_only = Class.objects.filter(academic_year=academic_year,school_id=school_id)
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(cls_name,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
        count =0
        class_section =[]
        cls_name_1 = []
        if request.method == "POST":
            class_name = request.POST.get('class_name')
            cls_name_1 =list(class_name.split(','))
            previous_cls = 0
            count = len(cls_name_1)
            for i in range(0,count):
                if '-' in cls_name_1[i] :
                    sec_split = list(cls_name_1[i].split('-'))
                    count_sec = len(sec_split)
                    new_cls = sec_split[0]
                    sec_name = sec_split[1]
                    existing_class = Class.objects.filter(class_name=new_cls,school_id=Scl_id,academic_year=academic_year)
                    if existing_class:
                        exist_id = Class.objects.only('class_id').get(class_name=new_cls,school_id=Scl_id,academic_year=academic_year).class_id
                        class_ex_id = Class.objects.get(class_id=exist_id)
                        section_exist = Section.objects.filter(section_name=sec_name,class_id=class_ex_id)
                        if section_exist:
                            pass
                        else:
                            sec_create = Section.objects.create(school_id=Scl_id,class_id=class_ex_id,section_name=sec_name)
                    else:
                        last_id = Class.objects.create(school_id=Scl_id,academic_year=academic_year,class_name=new_cls)
                        lst_id=last_id.class_id
                        class_last_id = Class.objects.get(class_id=int(lst_id))
                        section_exist = Section.objects.filter(section_name=sec_name,class_id=class_last_id)
                        if section_exist:
                            pass
                        else:
                            Section.objects.create(school_id=Scl_id,class_id=class_last_id,section_name=sec_name)
            if sec_create:
                messages.success(request,"Degree & Dept Added Successfully")
            if section_exist:
                messages.success(request,"Dept "+sec_name+ " Already Exist")
            return redirect("add_class_section")

        return render(request, 'add_classes_sections.html', {'cls_name':users,'class_id':class_id,'cls_only':cls_only,'cls_sec_name':cls_sec_name})
    else:
        return redirect("adminlogin")
def edit_classes_sections(request,cls_id,sec_id):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        if request.method == "POST":
            cls_name = request.POST.get('class_name')
            sec_name = request.POST.get('section_name')
            school_id = request.session['schoolname']
            academic_year = request.session['academic_year']
            cls_exist = Class.objects.filter(class_name=cls_name,school_id=school_id,academic_year=academic_year)
            section_exist = Section.objects.filter(class_id=cls_id,section_name=sec_name)
            if cls_exist:
                messages.success(request, "Class " + cls_name + " Already Exist")
            else:
                class_update = Class.objects.filter(class_id=cls_id).update(class_name=cls_name)
            if section_exist:
                messages.success(request, "Section " +sec_name + " Already Exist")
            else:
                section_update = Section.objects.filter(section_id=sec_id).update(section_name=sec_name)
            return redirect("add_class_section")
        post = Section.objects.filter(section_id=sec_id,class_id__in=Class.objects.filter(class_id=cls_id))
        return render(request, 'edit_classes_sections.html', {'post':post})
    else:
        return redirect("adminlogin")
def delete_classes_sections(request,cls_id,sec_id):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        section_del = Section.objects.get(section_id=sec_id)
        section_del.delete()
        return redirect("add_class_section")
    else:
        return redirect("adminlogin")
def edit_class(request,cls_id):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        post = Class.objects.filter(class_id=cls_id)
        school_id = request.session['schoolname']
        sch_id = SchoolDetail.objects.get(id=int(school_id))
        academic_year = request.session['academic_year']
        if request.method == 'POST':
            cls_name = request.POST.get('class_name')
            section_name = request.POST.get('section_name')
            sec_exist = Section.objects.filter(class_id=cls_id,section_name=section_name)
            class_id = Class.objects.get(class_id=cls_id)
            cls_exist = Class.objects.filter(class_name=cls_name,school_id=school_id,academic_year=academic_year)
            if cls_exist:
                messages.success(request, "Class " +cls_name + " Already Exist")
            else:
                Class.objects.filter(class_id=cls_id).update(class_name=cls_name)
            if sec_exist:
                messages.success(request, "Section " + section_name + "Already Exist")
            else:
                Section.objects.create(class_id=class_id,section_name=section_name,school_id=sch_id)
            return redirect("add_class_section")
        return render(request,"edit_class.html",{'post':post})
    else:
        return redirect("adminlogin")
def delete_class(request,cls_id):
	if request.session.has_key('username') and request.session.has_key('academic_year'):
		cls_del = Class.objects.get(class_id=cls_id)
		cls_del.delete()
		return redirect("add_class_section")
	else:
		return redirect("adminlogin")
def add_subject(request):
	if request.session.has_key('username') and request.session.has_key('academic_year'):
		sub = ""
		school_id = request.session['schoolname']
		academic_year = request.session['academic_year']
		row = Subject.objects.filter(school_id=school_id,class_id__in=Class.objects.filter(academic_year=academic_year)).select_related('class_id').select_related('section_id').order_by('-subject_id')
		page = request.GET.get('page', 1)
		page_count = pagination_cnt(request)
		paginator = Paginator(row,page_count)
		try:
			users = paginator.page(page)
		except PageNotAnInteger:
			users = paginator.page(1)
		except EmptyPage:
			users = paginator.page(paginator.num_pages)
		if request.method == "POST":
			get_subject = request.POST.get('subject_name')
			class_name = request.POST.get('class_id')
			exam_id = request.POST.get('exam_id')
			cls_id = Class.objects.get(class_id=int(class_name))
			sec_name = request.POST.get('section_id')
			sec_id = Section.objects.get(section_id=int(sec_name))
			school_name = SchoolDetail.objects.get(id=int(school_id))
			exam = Exam.objects.get(exam_id=int(exam_id))
			sub_name = list(get_subject.split(",")) 
			length = len(sub_name)
			for i in range(0,length):
				subject_exist = Subject.objects.filter(subject_name=sub_name[i],class_id=cls_id,section_id=sec_id,school_id=school_name,exam_id=exam)
				if subject_exist:
					messages.success(request, sub_name[i] + " Subject Already Exist")
				else:
					sub = Subject.objects.create(subject_name=sub_name[i],class_id=cls_id,section_id=sec_id,school_id=school_name,exam_id=exam)
					if sub:
						messages.success(request,"Subject Added Successfully")
			form = SubjectForm()
			return redirect("add_subject")
		else:
			form = SubjectForm()
			return render(request, 'add_subject.html', {'row':users,'form':form})
	else:
		return redirect("adminlogin")
def edit_subject(request,pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        post = get_object_or_404(Subject, pk=pk)
        if request.method == "POST":
            form = SubjectForm(request.POST,instance=post)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                return redirect('add_subject')
        else:
            form = SubjectForm(instance=post)
        return render(request, 'edit_subject.html', {'form': form})
    else:
        return redirect("adminlogin")
def delete_subject(request, pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        subject = Subject.objects.get(subject_id=pk)
        subject.delete()
        return redirect("add_subject")
    else:
        return redirect("adminlogin")
def assign_subjects_to_teachers(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        teacher = StaffDetail.objects.filter(school_id=school_id)
        class_id = Class.objects.filter(school_id=school_id,academic_year=academic_year)
        if request.method == "POST":
            form = AssignTeacherForm(request.POST)
            if form.is_valid():
                cls_id = request.POST.get('class_id')
                sec_id = request.POST.get('section_id')
                sub_id = request.POST.get('subject_id')
                staff_id = request.POST.get('staff_id')
                section_name1 = Section.objects.only('section_name').get(section_id=int(sec_id)).section_name
                class_id = Class.objects.get(class_id=int(cls_id))
                class_name = Class.objects.only('class_name').get(class_id=int(cls_id)).class_name
                section_id = Section.objects.get(section_id=int(sec_id))
                subject_id = Subject.objects.get(subject_id=int(sub_id))
                is_class_teacher = AssignSubjectTeacher.objects.filter(class_id=int(cls_id),section_id=int(sec_id),is_class_teacher='yes')
                subject_name = Subject.objects.only('subject_name').get(subject_id=int(sub_id)).subject_name
                teachet_exist = AssignSubjectTeacher.objects.filter(class_id=int(cls_id),section_id=int(sec_id),
                subject_id=int(sub_id))
                if teachet_exist:
                    messages.success(request,"Teacher Already Assigned For "+subject_name+ " Subject for " +class_name+ " -" +section_name1 + " Section")
                    return redirect("assign_teachers_view") 
                elif is_class_teacher:
                    messages.success(request,"Class Teacher Already Allocated For "+class_name+ " th Standard " +section_name1+ " Section")
                    return redirect("assign_teachers_view")  
                else:
                    post = form.save(commit=False)
                    post.academic_year = academic_year
                    post.save()
                    return redirect("assign_teachers_view")  
        else:
            form = AssignTeacherForm()
        return render(request, 'assign_subjects_to_teachers.html', {'form':form,'teacher':teacher,'class_id':class_id})
    else:
        return redirect("adminlogin")
def edit_assign_subjects_to_teachers(request,pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        post = get_object_or_404(AssignSubjectTeacher, pk=pk)
        teacher = AssignSubjectTeacher.objects.filter(assign_subject_teacher_id=pk)
        if request.method == "POST":
            form = AssignTeacherEditForm(request.POST,instance=post)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                return redirect("assign_teachers_view")
        else:
            form = AssignTeacherEditForm(instance=post)
        return render(request, 'edit_assign_subjects_to_teachers.html', {'form':form,'teacher':teacher})
    else:
        return redirect("adminlogin")
def delete_assign_teacher(request,pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        subject = AssignSubjectTeacher.objects.get(assign_subject_teacher_id=pk)
        subject.delete()
        return redirect("assign_teachers_view")
    else:
        return redirect("adminlogin")
def add_exam(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        ex = ""
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        exam = Exam.objects.filter(school_id=school_id ,class_id__in=Class.objects.filter(academic_year=academic_year)).select_related('class_id').select_related('section_id').order_by('-exam_id')
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(exam,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
        if request.method == "POST":
            get_exam = request.POST.get('exams')
            class_name = request.POST.get('class_id')
            cls_id = Class.objects.get(class_id=int(class_name))
            sec_name = request.POST.get('section_id')
            sec_id = Section.objects.get(section_id=int(sec_name))
            school_name = SchoolDetail.objects.get(id=int(school_id))
            exam_name = list(get_exam.split(",")) 
            length = len(exam_name)
            for i in range(0,length):
                exam_exist = Exam.objects.filter(exams=exam_name[i],class_id=cls_id,section_id=sec_id,school_id=school_name)
                if exam_exist:
                    messages.success(request, exam_name[i] + " Already Exist")
                else:
                    ex = Exam.objects.create(exams=exam_name[i],class_id=cls_id,section_id=sec_id,school_id=school_name)
            if ex:
                messages.success(request,"Exam Added Successfully")
            form = ExamForm()
        else:
            form = ExamForm()
        return render(request, 'add_exam.html', {'form':form,'exam':users})
    else:
        return redirect("adminlogin")
def edit_exam(request,pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        post = get_object_or_404(Exam, pk=pk)
        if request.method == "POST":
            form = ExamEditForm(request.POST,instance=post)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                return redirect('add_exam')
        else:
            form = ExamEditForm(instance=post)
        return render(request, 'edit_exam.html', {'form': form})
    else:
        return redirect("adminlogin")
def delete_exam(request, pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        exam = Exam.objects.get(exam_id=pk)
        exam.delete()
        return redirect("add_exam")
    else:
        return redirect("adminlogin")
def pagination_count_year(request):
    if request.session.has_key('teacher'):
        school_id = request.session['schoolname']
        page_count = SchoolSetting.objects.filter(school_id=school_id)
        if request.method == "POST":
            del request.session['academic_year']
            settings_id = request.POST.get('settings_id')
            count = request.POST.get('page_count')
            cnt = int(count)
            academic_year = request.POST.get('academic_year')
            set_id = int(settings_id)
            SchoolSetting.objects.filter(settings_id=set_id).update(academic_year=academic_year,pagination_count=cnt)
            messages.success(request,"Settings Updated Successfully")
        return render(request, 'other_settings.html', {'page_count':page_count})
    else:
        return redirect("teacher_login")
def pagination_count(request):
    if request.session.has_key('username'):
        school_id = request.session['schoolname']
        page_count = SchoolSetting.objects.filter(school_id=school_id)
        if request.method == "POST":
            del request.session['academic_year']
            settings_id = request.POST.get('settings_id')
            count = request.POST.get('page_count')
            cnt = int(count)
            academic_year = request.POST.get('academic_year')
            set_id = int(settings_id)
            SchoolSetting.objects.filter(settings_id=set_id).update(academic_year=academic_year,pagination_count=cnt)
            messages.success(request,"Settings Updated Successfully")
        return render(request, 'admin_settings.html', {'page_count':page_count})
    else:
        return redirect("adminlogin")
def student_section(request):
	if request.session.has_key('username') and request.session.has_key('academic_year'):
		school_id=request.session['schoolname']
		academic_year = request.session['academic_year']
		class_name = Class.objects.filter(school_id=school_id,academic_year=academic_year)
		student = StudentDetail.objects.filter(school_id=school_id)
		if "GET" == request.method:
			return render(request, 'student_section.html', {'student':student,'cls_name':class_name})
		else:
			file = request.FILES['excel_file']
			school_id = request.session['schoolname']
			ids = SchoolDetail.objects.get(id=int(school_id))
			class_id = request.POST.get('class_id')
			student_class = Class.objects.get(class_id=int(class_id)) 
			section_id = request.POST.get('section_id')
			student_section = Section.objects.get(section_id=int(section_id))
			workbook = openpyxl.load_workbook(file, read_only=True)
			# Get name of the first sheet and then open sheet by name
			first_sheet = workbook.get_sheet_names()[0]
			worksheet = workbook.get_sheet_by_name(first_sheet)

			data = []
			for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
				register_num = row[0].value
				student_name = row[1].value
				filter_student = StudentDetail.objects.filter(register_number=register_num,school_id=ids)
				if filter_student:
					get_student_id = StudentDetail.objects.only('id').get(register_number=register_num).id
					stud_id = StudentDetail.objects.get(id=int(get_student_id))
					user_exist = StudentSection.objects.filter(student_id=stud_id,academic_year=academic_year)

					if user_exist:
						messages.success(request, student_name + " Class Already Assigned")
					else:
						student = StudentSection()
						student.student_id = stud_id
						student.class_id = student_class
						student.section_id = student_section
						student.academic_year = academic_year
						student.reg_no = register_num
						student.student_name = student_name
						data.append(student)
				else:
					messages.success(request, "Please Check the " + student_name + "'s Register Number is Incorrect.")
				# Bulk create data
			StudentSection.objects.bulk_create(data)
			return redirect("view_student_section")

		return render(request, 'student_section.html', {'student':student,'cls_name':class_name})
	else:
		return redirect("adminlogin")
def edit_student_section(request,pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        post = get_object_or_404(StudentSection, pk=pk)
        if request.method == "POST":
            form = StudentSectionEditForm(request.POST,instance=post)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                return redirect("view_student_section")
        else:
            form = StudentSectionEditForm(instance=post)
        return render(request, 'edit_student_section.html', {'form': form , 'post' :post})
    else:
        return redirect("adminlogin")
def delete_student_section(request, pk):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        student = StudentSection.objects.get(student_section_id=pk)
        student.delete()
        return redirect("view_student_section")
    else:
        return redirect("adminlogin")
def assign_student_class(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        if request.GET.get('cls_id') and request.GET.get('sec_id'):
            cls_id = request.GET.get('cls_id')
            sec_id = request.GET.get('sec_id')
            student_class = Class.objects.get(class_id=int(cls_id))
            student_section = Section.objects.get(section_id=int(sec_id))
            school_id = request.session['schoolname']
            ids = SchoolDetail.objects.get(id=int(school_id))
            academic_year = request.session['academic_year']
            if "GET" == request.method:
                return render(request, 'assign_student_class.html', {})
            else:
                file = request.FILES['excel_file']
                workbook = openpyxl.load_workbook(file, read_only=True)
                # Get name of the first sheet and then open sheet by name
                first_sheet = workbook.get_sheet_names()[0]
                worksheet = workbook.get_sheet_by_name(first_sheet)

                data = []
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
                    register_num = row[0].value
                    student_name = row[1].value
                    filter_student = StudentDetail.objects.filter(register_number=register_num,school_id=ids)
                    if filter_student:
                        get_student_id = StudentDetail.objects.only('id').get(register_number=register_num).id
                        stud_id = StudentDetail.objects.get(id=int(get_student_id))
                        user_exist = StudentSection.objects.filter(student_id=stud_id,academic_year=academic_year)

                        if user_exist:
                            messages.success(request, student_name + " Class Already Assigned")
                        else:
                            student = StudentSection()
                            student.student_id = stud_id
                            student.class_id = student_class
                            student.section_id = student_section
                            student.academic_year = academic_year
                            data.append(student)
                    else:
                        messages.success(request, "Please Check the " + student_name + "'s Register Number is Incorrect.")
                # Bulk create data
                create_student = StudentSection.objects.bulk_create(data)
                if create_student:
                    messages.success(request,"Students Class & Section Assigned Successfully")
            return render(request,"assign_student_class.html",{})
        else:
            return redirect("class_details")
    else:
        return redirect("teacher_login")
def teacher_edit_student_section(request,pk):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        post = get_object_or_404(StudentSection, pk=pk)
        if request.method == "POST":
            form = StudentSectionEditForm(request.POST,instance=post)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                messages.success(request,"Updated Successfully")
        else:
            form = StudentSectionEditForm(instance=post)
        return render(request, 'teacher_edit_student_section.html', {'form': form , 'post' :post})
    else:
        return redirect("teacher_login")
def teacher_delete_student_section(request, pk):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        student = StudentSection.objects.get(student_section_id=pk)
        student.delete()
        return redirect("class_details")
    else:
        return redirect("teacher_login")
def search_student_year(request):
	if request.session.has_key('teacher') and request.session.has_key('academic_year'):
		school_id = request.session['schoolname']
		years = Class.objects.filter(school_id=school_id).order_by('-academic_year')
		years = years.values('academic_year').annotate(count=Count('academic_year'))
		return render(request, 'search_student_year.html', {'years':years})
	else:
		return redirect("teacher_login")

def ajax_exam_class_year_search(request):
    academic_year = request.POST.get('academic_year')
    school_id = request.session['schoolname']
    teacher_id = request.session['user_id']
    cursor = connection.cursor()
    post = '''SELECT school_class.class_id,school_class.class_name from school_class INNER JOIN school_assignsubjectteacher
    ON school_class.class_id=school_assignsubjectteacher.class_id_id where school_class.academic_year='%s' AND school_class.school_id_id='%d'
    AND school_assignsubjectteacher.staff_id_id='%d' Group By school_class.class_id''' % (academic_year,school_id,teacher_id)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data = {
    'class_name':row
    }
    return JsonResponse(data)
def ajax_exam_class_select_section(request):
    school_id = request.session['schoolname']
    class_id = request.POST.get('class_id')
    academic_year = request.session['academic_year']
    teacher_id = request.session['user_id']
    cursor = connection.cursor()
    post = '''SELECT school_section.section_id,school_section.section_name from school_class INNER JOIN school_section 
    ON school_class.class_id=school_section.class_id_id INNER JOIN school_assignsubjectteacher
    ON school_section.section_id=school_assignsubjectteacher.section_id_id where school_section.class_id_id='%d'
    AND school_class.school_id_id='%d' AND school_class.academic_year='%s' AND school_section.school_id_id='%d' 
    AND school_assignsubjectteacher.staff_id_id='%d' Group By school_section.section_id ''' % (int(class_id),int(school_id),academic_year,int(school_id),teacher_id)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data =  {

        'rowval':row
    }
    return JsonResponse(data)
def student_mark(request):
	if request.session.has_key('username') and request.session.has_key('academic_year'):
		school_id = request.session['schoolname']
		years = Class.objects.filter(school_id=school_id).order_by('-academic_year')
		years = years.values('academic_year').annotate(count=Count('academic_year'))
		return render(request, 'student_mark.html', {'years':years})
	else:
		return redirect("adminlogin")

def ajax_exam_class_search(request):
    academic_year = request.POST.get('academic_year')
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    post = '''SELECT app_class.class_id,app_class.class_name from app_class where app_class.academic_year='%s' AND app_class.school_id_id='%d' ''' % (academic_year,school_id)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data = {
    'class_name':row
    }
    return JsonResponse(data)


def student_view_marks(request):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
        student_id = request.session['student_id']
        student_name = StudentSection.objects.filter(student_id=student_id)
        return render(request, 'student_view_marks.html', {'student_name':student_name})
    else:
        return redirect("student_login")
def student_view_diary_notes(request):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
        student_id = request.session['student_id']
        cursor = connection.cursor()
        student = ''' SELECT * from school_studentsection where school_studentsection.student_id_id='%d' ''' % (student_id)
        query = cursor.execute(student)
        row = cursor.fetchone()
        cls_id = row[2]
        sec_id = row[3]
        task = StudentDiary.objects.filter(class_id=cls_id,section_id=sec_id)
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(task,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
        return render(request, 'student_view_diary_notes.html', {'student':users})
    else:
        return redirect("student_login")
def manage_student_marks(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        school_id=request.session['schoolname']
        academic_year=request.session['academic_year']
        staff_id=request.session['user_id']
        staff = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,class_id__in=Class.objects.filter(academic_year=academic_year,school_id_id=school_id)).select_related('class_id').select_related('section_id').select_related('subject_id')
        mark=Exam.objects.filter(school_id_id=school_id,class_id__in=Class.objects.filter(academic_year=academic_year)).select_related('class_id')   
        class_id = staff.values('class_id')
        section_id = staff.values('section_id')

        student = StudentSection.objects.filter(class_id__in=AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,class_id__in=Class.objects.filter(
        academic_year=academic_year,school_id_id=school_id)).select_related('class_id').select_related('section_id').values('class_id'),section_id__in=AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,class_id__in=Class.objects.filter(
        academic_year=academic_year,school_id_id=school_id)).select_related('class_id').select_related('section_id').values('section_id'))

        return render(request, 'manage_student_marks.html', {'mark':mark,'staff':staff,'class_id':class_id,'section_id':section_id,'student':student})
    else:
        return redirect("teacher_login")
def view_manage_student_marks(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        if request.session['user_id']:
            cls_id = request.GET.get('cls_id')
            sec_id = request.GET.get('sec_id')
            class_id = int(cls_id)
            section_id = int(sec_id)
            school_id = request.session['schoolname']
            staff_id = request.session['user_id']
            academic_year = request.session['academic_year']
            student = StudentSection.objects.filter(class_id_id=class_id,
            section_id_id=section_id,academic_year=academic_year,student_id__in=StudentDetail.objects.filter(school_id=school_id)) 
            return render(request, 'view_manage_student_marks.html', {'post':student})
        else:
            return redirect("teacher_login")
    else:
        return redirect("teacher_login")
def add_student_diary_notes(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        if request.GET.get('class_id') and request.GET.get('sec_id'):
            class_id = request.GET.get('class_id')
            cls_id = int(class_id)
            section_id = request.GET.get('sec_id')
            sec_id = int(section_id)
            staff_id=request.session['user_id']
            if request.method == "POST":
                homework_classwork = int(request.POST.get('homework_classwork'))
                subject = request.POST.get('subject_id_id')
                assigned_date = request.POST.get('assigned_date')
                assigned_date =datetime.strptime(assigned_date, "%d-%m-%Y").strftime('%Y-%m-%d')
                diary_notes = request.POST.get('diary_notes')
                StudentDiary.objects.create(subject_id_id=subject,assigned_date=assigned_date, diary_note=diary_notes,class_id_id=cls_id,section_id_id=sec_id,staff_id_id=staff_id, homework_classwork=homework_classwork)
                messages.success(request,'Diary Added Successfully')
            school_id = request.session['schoolname']
            academic_year = request.session['academic_year']
            student = StudentSection.objects.filter(class_id=cls_id,academic_year=academic_year,section_id=sec_id,student_id_id__in=StudentDetail.objects.filter(school_id=school_id) )
            subject = AssignSubjectTeacher.objects.filter(section_id=sec_id,class_id__in=Class.objects.filter(academic_year=academic_year,class_id=int(cls_id)), staff_id=staff_id)
            return render(request,'add_student_diary_notes.html',{'subject':subject,'student':student})
        else:
            return redirect("teacher_class_diary")
    else:
        return redirect("teacher_login")
def view_student_diary_notes(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        diary_data = dict()
        if request.method == "POST":
            diary_id = request.POST.get('diary_id')
            ids = int(diary_id)
            diary = StudentDiary.objects.get(diary_id=ids)
            diary.delete()
        if request.GET.get('class_id') and request.GET.get('sec_id'):
            class_id = request.GET.get('class_id')
            cls_id = int(class_id)
            section_id = request.GET.get('sec_id')
            sec_id = int(section_id)
            staff_id=request.session['user_id']
            school_id = request.session['schoolname']
            academic_year = request.session['academic_year']
            diary_data = StudentDiary.objects.filter(class_id_id__in=Class.objects.filter(class_id=cls_id,academic_year=academic_year),section_id_id=sec_id,staff_id_id=staff_id )
            page = request.GET.get('page', 1)
            page_count = pagination_cnt(request)
            paginator = Paginator(diary_data,page_count)
            try:
                users = paginator.page(page)
            except PageNotAnInteger:
                users = paginator.page(1)
            except EmptyPage:
                users = paginator.page(paginator.num_pages)
        return render(request,'view_student_diary_notes.html',{'diary_data':users})
    else:
        return redirect("teacher_login")
def teacher_change_password(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        staff_id=request.session['user_id']
        get_teacher=""
        if request.method == "POST":
            get_psw =  request.POST['password']
            password = StaffDetail.objects.filter(staff_id=staff_id).update(password=get_psw)
            messages.success(request, 'Password Updated Successfully')
        else:
            get_teacher = StaffDetail.objects.filter(staff_id=staff_id)
        return render(request,'teacher_change_password.html',{"teacher": get_teacher})
    else:
        return redirect("teacher_login")

def view_student_section(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        academic_year = request.session['academic_year'];
        school_id = request.session['schoolname'];
        row = StudentSection.objects.filter(academic_year=academic_year,student_id__in=StudentDetail.objects.filter(school_id_id=school_id)).select_related('student_id').select_related('class_id').select_related('section_id').order_by('-student_section_id')
        length = len(row)
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(row,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
        return render(request, 'view_student_section.html', {'row' : users,'length':length})
    else:
        return redirect("adminlogin")
def teacher_view_student_section(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        if request.GET.get('cls_id') and request.GET.get('sec_id'):
            academic_year = request.session['academic_year']
            school_id = request.session['schoolname']
            teacher_id = request.session['user_id']
            class_id = request.GET.get('cls_id')
            section_id = request.GET.get('sec_id')
            row = StudentSection.objects.filter(academic_year=academic_year,student_id__in=StudentDetail.objects.filter(school_id_id=school_id),
            class_id=int(class_id),section_id=int(section_id)).select_related('student_id').select_related('class_id').select_related('section_id').order_by('-student_section_id')
            length = len(row)
            page = request.GET.get('page', 1)
            page_count = pagination_cnt(request)
            paginator = Paginator(row,page_count)
            try:
                users = paginator.page(page)
            except PageNotAnInteger:
                users = paginator.page(1)
            except EmptyPage:
                users = paginator.page(paginator.num_pages)
            return render(request, 'teacher_view_student_section.html', {'row' : users,'length':length})
        else:
            return redirect("class_details")
    else:
        return redirect("teacher_login")

def select_section(request):
    school_id = request.session['schoolname']
    class_id = request.POST.get('class_id')
    academic_year = request.session['academic_year']
    cursor = connection.cursor()
    post = '''SELECT * from app_class INNER JOIN app_section ON
    app_class.class_id=app_section.class_id_id where app_section.class_id_id='%d'
    AND app_class.school_id_id='%d' AND app_class.academic_year='%s' AND app_section.school_id_id='%d' ''' % (int(class_id),int(school_id),academic_year,int(school_id))
    query = cursor.execute(post)
    row = cursor.fetchall()
    data =  {

        'rowval':row
    }
    return JsonResponse(data)
def select_exam_college(request):
    school_id = request.session['schoolname']
    class_id = request.POST.get('section_id')
    academic_year = request.session['academic_year']
    cursor = connection.cursor()
    post = '''SELECT app_exam.exam_id,app_exam.exams from app_section INNER JOIN app_exam ON
    app_section.section_id=app_exam.section_id_id INNER JOIN app_class ON app_exam.class_id_id=app_class.class_id where app_exam.section_id_id='%d'
    AND app_exam.school_id_id='%d' AND app_class.academic_year='%s' ''' % (int(class_id),int(school_id),academic_year)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data =  {

        'rowval':row
    }
    return JsonResponse(data)

def select_teacher_section(request):
    school_id = request.session['schoolname']
    class_id = request.POST.get('class_id')
    academic_year = request.session['academic_year']
    teacher_id = request.session['user_id']
    cursor = connection.cursor()
    post = '''SELECT COUNT(school_class.class_id),COUNT(school_section.section_id), from school_class INNER JOIN school_section ON
    school_class.class_id=school_section.class_id_id INNER JOIN school_assignsubjectteacher ON
    school_assignsubjectteacher.section_id_id=school_section.section_id  where school_section.class_id_id='%d'
    AND school_class.school_id_id='%d' AND school_class.academic_year='%s' AND school_section.school_id_id='%d'
    AND school_assignsubjectteacher.staff_id_id='%d' GROUP BY  school_class.class_id''' % (int(class_id),int(school_id),academic_year,int(school_id),int(teacher_id))
    query = cursor.execute(post)
    row = cursor.fetchall()
    data =  {

        'rowval':row
    }
    return JsonResponse(data)

def select_academic_year(request):
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    post = '''SELECT * from app_schoolsetting where app_schoolsetting.school_id_id='%d' ''' %(int(school_id))
    query = cursor.execute(post)
    row = cursor.fetchone()
    year = row[1]
    request.session['academic_year'] = year
    academic_year =  request.session['academic_year']
    data = {

    'academic_year':academic_year

    }
    return JsonResponse(data)

def select_class(request):
    school_id = request.POST.get('school_id')
    academic_year = request.session['academic_year']
    cursor = connection.cursor()
    post = '''SELECT * from app_class where app_class.academic_year='%s' ''' % (academic_year)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data =  {

        'row':row
    }
    return JsonResponse(data)

def student_section_class_select(request):
    school_id = request.session['schoolname']
    class_id = request.POST.get('class_id')
    academic_year = request.session['academic_year']

    cursor = connection.cursor()
    post = '''SELECT * from app_class INNER JOIN app_section ON
    app_class.class_id=app_section.class_id_id where app_section.class_id_id='%d'
    AND app_class.school_id_id='%d' AND app_class.academic_year='%s' ''' % (int(class_id),int(school_id),academic_year)
    query = cursor.execute(post)
    row = cursor.fetchall()

    data = {
        'row':row
    }

    return JsonResponse(data)

def section_selection_student(request):

    sec_id = request.GET.get('id')
    cursor = connection.cursor()
    post = '''SELECT * from school_StudentSection  where school_StudentSection.student_section_id='%d'
     ''' % (int(sec_id))
    query = cursor.execute(post)
    row = cursor.fetchall()

    data = {
        'row':row
    }
    return JsonResponse(data)

def select_subject(request):
    ids = request.POST.get('class_id')
    class_id = int(ids)
    sec_id = request.POST.get('section_id')
    section_id = int(sec_id)
    academic_year = request.session['academic_year']
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    post = '''SELECT * from school_class INNER JOIN school_section ON school_class.class_id=school_section.class_id_id
    INNER JOIN school_subject ON school_subject.section_id_id=school_section.section_id where school_subject.class_id_id ='%d'
    AND school_subject.section_id_id='%d' AND school_class.school_id_id='%d' AND school_class.academic_year='%s' AND school_section.school_id_id='%d' ''' % (class_id,section_id,school_id,academic_year,school_id)
    query = cursor.execute(post)
    row = cursor.fetchall()

    data = {
        'subject':row
    }
    return JsonResponse(data)

def assign_teachers_view(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        academic_year = request.session['academic_year']
        school_id = request.session['schoolname']
        teachers = AssignSubjectTeacher.objects.filter(academic_year=academic_year,staff_id__in=StaffDetail.objects.filter(school_id_id=school_id)).select_related('class_id').select_related('section_id').select_related('subject_id').select_related('staff_id').order_by('-assign_subject_teacher_id')
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(teachers,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
        return render(request,'assign_teachers_view.html',{'teachers':users})
    else:
        return redirect("adminlogin")
def admin_change_password(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        admin_id = request.session['schoolname']
        get_admin=""
        if request.method == "POST":
            get_psw =  request.POST['password']
            password = SchoolDetail.objects.filter(id=admin_id).update(password=get_psw)
            messages.success(request, 'Password Updated Successfully')
        else:
            get_admin = SchoolDetail.objects.filter(id=admin_id)
        return render(request,'change_password.html',{"school": get_admin})
    else:
        return redirect("adminlogin")
def manage_students_marks(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        ex=""
        school_id=request.session['schoolname']
        academic_year=request.session['academic_year']
        staff_id=request.session['user_id']
        query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year)
        staff = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('class_id')).order_by('class_id__class_name')
        class_id = query.values('class_id')
        section_id = query.values('section_id')
        count = len(class_id)
        for i in range(0,count):
        	exams = Exam.objects.filter(class_id__in=class_id,section_id__in=section_id)
        	ex = exams.values('class_id', 'section_id','class_id__class_name','section_id__section_name','exam_id','exams').annotate(count = Count('class_id'))
        
        return render(request,'manage_students_marks.html',{'staff':staff,'class_id':class_id,'exams':ex,'section_id':section_id,'query':query})
    else:
        return redirect("teacher_login")
def is_class_teacher(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        school_id=request.session['schoolname']
        academic_year=request.session['academic_year']
        staff_id=request.session['user_id']
        query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year,is_class_teacher='yes')
        staff = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('class_id')).order_by('class_id__class_name')
        return render(request,'is_class_teacher.html',{'staff':staff})
    else:
        return redirect("teacher_login")
def class_details(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        school_id=request.session['schoolname']
        academic_year=request.session['academic_year']
        staff_id=request.session['user_id']
        query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year,is_class_teacher='yes')
        staff = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('class_id')).order_by('class_id__class_name')
        return render(request,'class_details.html',{'staff':staff})
    else:
        return redirect("teacher_login")
def class_student_mark_details(request,class_id,section_id,exam_id):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        cursor = connection.cursor()
        post = '''SELECT school_mark.student_reg_no, SUM(school_mark.mark), Avg(school_mark.mark) from school_mark INNER JOIN school_studentdetail ON school_mark.student_reg_no=school_studentdetail.register_number
        where school_mark.class_id ='%d'AND school_mark.section_id='%d' AND school_mark.exam_id='%d' GROUP BY school_mark.student_reg_no order by school_mark.student_reg_no''' % (class_id,section_id,exam_id)
        ex = cursor.execute(post)
        row = cursor.fetchall()
        sub1 = '''SELECT  COUNT(school_subject.subject_id),school_subject.subject_name,school_subject.subject_id from school_mark INNER JOIN school_subject ON school_mark.subject_id=school_subject.subject_id
        where school_mark.class_id ='%d'AND school_mark.section_id='%d'  GROUP BY school_subject.subject_id order by school_subject.subject_id''' % (class_id,section_id)
        ex1 = cursor.execute(sub1)
        sub_name = cursor.fetchall()
        cnt  = len(sub_name)
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(row,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)
        return render(request,'class_student_mark_details.html',{'row':users,'class_id':class_id,
        'section_id':section_id,'exam_id':exam_id,'sub_name':sub_name,'tot':cnt})
    else:
        return redirect("teacher_login")
def select_class_exam(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        if request.GET.get('cls_id') and request.GET.get('sec_id'):
            cls_id = request.GET.get('cls_id')
            sec_id = request.GET.get('sec_id')
            class_id = int(cls_id)
            section_id = int(sec_id)
            school_id = request.session['schoolname']
            exams = Exam.objects.filter(class_id_id=class_id,section_id_id=section_id,school_id_id=school_id)
            return render(request,'select_class_exam.html',{'exams':exams})
        else:
            return redirect("is_class_teacher")
    else:
        return redirect("teacher_login")
def ajax_student_marks(request):
    cls_id = request.POST.get('class_id')
    sec_id = request.POST.get('section_id')
    sub_id = request.POST.get('subject_id')
    request.session['class_id'] = cls_id
    data = {
    'cls_id':cls_id,
    'sec_id':sec_id,
    'sub_id':sub_id
    }
    return JsonResponse(data)
def add_student_marks(request,cls_id,sec_id,exam_id):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        rc = dict()
        st = StudentSection.objects.filter(class_id=cls_id,section_id=sec_id)
        cnt = len(st)
        teacher_id = request.session['user_id']
        cursor = connection.cursor()
        get_mark = ''' SELECT school_mark.*,school_studentdetail.register_number,school_studentdetail.student_name,school_subject.subject_name,school_class.class_name,school_section.section_name from school_mark
                        LEFT JOIN school_studentdetail ON school_studentdetail.register_number=school_mark.student_reg_no
                        LEFT JOIN school_subject ON school_subject.subject_id=school_mark.subject_id
                        LEFT JOIN school_class ON school_class.class_id=school_mark.class_id
                        LEFT JOIN school_section ON school_section.section_id=school_mark.section_id
                        WHERE school_mark.class_id=%d AND school_mark.section_id=%d AND school_mark.staff_id=%d AND school_mark.exam_id=%d '''%(int(cls_id),int(sec_id),int(teacher_id),int(exam_id))
        sub = cursor.execute(get_mark)
        mark = cursor.fetchall()
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']

        if request.method == "POST":
            if request.POST.getlist('mark_id[]'):
                ids = request.POST.getlist('mark_id[]')
                mark = request.POST.getlist('student_mark[]')
                length=len(mark)
                for row in range(0,length):
                    if mark[row]=="" :
                        mark[row]=0;
                    Mark.objects.filter(mark_id=ids[row]).update(mark=mark[row])
                return redirect("manage_students_marks")
            else :
                subject_id = request.POST.getlist('subject_id[]')
                mark = request.POST.getlist('student_mark[]')
                student_id = request.POST.getlist('stud_reg_no[]')
                staff_id = request.POST.getlist('staff_id[]')
                length=len(mark)
                for row in range(0,length):
                    if mark[row]=="":
                        mark[row]=0
                    Mark.objects.create(student_reg_no=student_id[row],class_id=cls_id,section_id=sec_id,subject_id=subject_id[row],staff_id=staff_id[row],mark=mark[row],exam_id=exam_id)
                return redirect("manage_students_marks")
        #exam = Exam.objects.filter(section_id=int(sec_id),school_id=school_id,class_id__in=Class.objects.filter(academic_year=academic_year,class_id=int(cls_id)))
        exam = Exam.objects.filter(exam_id=exam_id)
        student = StudentSection.objects.filter(class_id=cls_id,academic_year=academic_year,section_id=sec_id,student_id_id__in=StudentDetail.objects.filter(school_id=school_id))
        subject = AssignSubjectTeacher.objects.filter(section_id=sec_id,class_id__in=Class.objects.filter(academic_year=academic_year,class_id=int(cls_id)),staff_id=teacher_id)
            
        return render(request,'add_student_marks.html',{'exam':exam,'student':student,'subject':subject,'student_reg':rc, 'mark' :mark,'cnt':cnt})
    else:
        return redirect("teacher_login")
def import_mark(request,class_id,section_id,exam_id):
    mark_exsit = ""
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        st = StudentSection.objects.filter(class_id=class_id,section_id=section_id)
        cnt = len(st)
        result = ""
        staff_id = request.session['user_id']
        clss_id = Section.objects.filter(section_id=section_id).select_related('class_id')
        val = clss_id.values('class_id_id','section_id')
        ex_name = Exam.objects.filter(exam_id=exam_id)
        subject_name = AssignSubjectTeacher.objects.filter(class_id=class_id,section_id=section_id,staff_id=staff_id).order_by('subject_id')
        if "GET" == request.method:
            return render(request, 'import_mark.html', {'subject_name':subject_name,'ex_name':ex_name,'val':val,'cnt':cnt})
        else:
            excel_file = request.FILES["excel_file"]

            wb = openpyxl.load_workbook(excel_file)

            worksheet = wb["Mark Sheet"]

            mylist = []
            for cell in worksheet['C']:
                if cell.value != None:
                    mylist.append(cell.value)

            mark1 = []
            for cell in worksheet['D']:
                if cell.value != None:
                    mark1.append(cell.value)
            m1 = len(mark1)

            mark2 = []
            for cell in worksheet['E']:
                if cell.value != None:
                    mark2.append(cell.value)
            m2 = len(mark2)

            mark3 = []
            for cell in worksheet['F']:
                if cell.value != None:
                    mark3.append(cell.value)
            m3 = len(mark3)

            mark4 = []
            for cell in worksheet['G']:
                if cell.value != None:
                    mark4.append(cell.value)
            m4 = len(mark4)

            mark5 = []
            for cell in worksheet['H']:
                if cell.value != None:
                    mark5.append(cell.value)
            m5 = len(mark5)

            mark6 = []
            for cell in worksheet['I']:
                if cell.value != None:
                    mark6.append(cell.value)
            m6 = len(mark6)

            mark7 = []
            for cell in worksheet['J']:
                if cell.value != None:
                    mark7.append(cell.value)
            m7 = len(mark7)

            mark8 = []
            for cell in worksheet['K']:
                if cell.value != None:
                    mark8.append(cell.value)
            m8 = len(mark8)

            mark9 = []
            for cell in worksheet['L']:
                if cell.value != None:
                    mark9.append(cell.value)
            m9 = len(mark9)

            mark10 = []
            for cell in worksheet['M']:
                if cell.value != None:
                    mark10.append(cell.value)
            m10 = len(mark10)


            cls_id = class_id
            sec_id = section_id
            ex_id = exam_id

            staff_id = request.session['user_id']
            sub_id = request.POST.getlist('sub_id[]')
            lenthsub = len(sub_id)
            
            if lenthsub == 1:
                sub1 = sub_id[0]
            if lenthsub == 2:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
            if lenthsub == 3:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
                sub3 = sub_id[2]
            if lenthsub == 4:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
                sub3 = sub_id[2]
                sub4 = sub_id[3]
            if lenthsub == 5:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
                sub3 = sub_id[2]
                sub4 = sub_id[3]
                sub5 = sub_id[4]
            if lenthsub == 6:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
                sub3 = sub_id[2]
                sub4 = sub_id[3]
                sub5 = sub_id[4]
                sub6 = sub_id[5]
            if lenthsub == 7:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
                sub3 = sub_id[2]
                sub4 = sub_id[3]
                sub5 = sub_id[4]
                sub6 = sub_id[5]
                sub7 = sub_id[6]
            if lenthsub == 8:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
                sub3 = sub_id[2]
                sub4 = sub_id[3]
                sub5 = sub_id[4]
                sub6 = sub_id[5]
                sub7 = sub_id[6]
                sub8 = sub_id[7]
            if lenthsub == 9:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
                sub3 = sub_id[2]
                sub4 = sub_id[3]
                sub5 = sub_id[4]
                sub6 = sub_id[5]
                sub7 = sub_id[6]
                sub8 = sub_id[7]
                sub9 = sub_id[8]
            if lenthsub == 10:
                sub1 = sub_id[0]
                sub2 = sub_id[1]
                sub3 = sub_id[2]
                sub4 = sub_id[3]
                sub5 = sub_id[4]
                sub6 = sub_id[5]
                sub7 = sub_id[6]
                sub8 = sub_id[7]
                sub9 = sub_id[8]
                sub10 = sub_id[9]

            length=len(mylist)

            for row in range(1,length):
                if m1 > 1 and sub1!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub1)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub1).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub1).update(mark=mark1[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub1,mark=mark1[row])
                if m2 > 1 and sub2!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub2)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub2).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub2).update(mark=mark2[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub2,mark=mark2[row])
                if m3 > 1 and sub3!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub3)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub2).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub3).update(mark=mark3[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub3,mark=mark3[row])
                if m4 > 1 and sub4!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub4)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub4).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub4).update(mark=mark4[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub4,mark=mark4[row])
                if m5 > 1 and sub5!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub5)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub5).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub5).update(mark=mark5[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub5,mark=mark5[row])
                if m6 > 1 and sub6!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub6)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub6).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub6).update(mark=mark6[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub6,mark=mark6[row])
                if m7 > 1 and sub7!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub7)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub7).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub7).update(mark=mark7[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub7,mark=mark7[row])
                if m8 > 1 and sub8!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub8)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub8).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub8).update(mark=mark8[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub8,mark=mark8[row])
                if m9 > 1 and sub9!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub9)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub9).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub9).update(mark=mark9[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub9,mark=mark9[row])
                if m10 > 1 and sub10!="":
                    mark_exsit = Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub10)
                    sub_name = Subject.objects.only('subject_name').get(subject_id=sub10).subject_name
                    if mark_exsit:
                        Mark.objects.filter(student_reg_no=mylist[row],exam_id=exam_id,subject_id=sub10).update(mark=mark10[row])
                    else:
                        result = Mark.objects.create(student_reg_no=mylist[row],class_id=cls_id,section_id=sec_id,staff_id=staff_id,exam_id=exam_id,subject_id=sub10,mark=mark10[row])
            if result:
                messages.success(request, "Mark Added Successfully")
            if mark_exsit:
                messages.success(request, "Mark Added Successfully")

            return render(request, 'import_mark.html', {'subject_name':subject_name,'ex_name':ex_name,'val':val,'cnt':cnt})
    else:
        return redirect("teacher_login")


def ajax_student_grap(request):
    cls_id = request.POST.get('cls_id')
    sec_id = request.POST.get('sec_id')
    class_id = int(cls_id)
    section_id = int(sec_id)
    school_id = request.session['schoolname']
    staff_id = request.session['user_id']
    academic_year = request.session['academic_year']
    query = StudentSection.objects.filter(class_id_id=class_id,section_id_id=section_id,
    academic_year=academic_year)
    student_count = query.values('student_id_id').aggregate(count = Count('student_id_id'))
    cursor = connection.cursor()
    subject = '''SELECT COUNT(school_mark.subject_id),SUM(school_mark.mark),school_subject.subject_name,COUNT(school_mark.student_reg_no),school_subject.subject_id from school_class INNER JOIN school_mark ON school_class.class_id=school_mark.class_id
    INNER JOIN school_subject ON school_subject.subject_id=school_mark.subject_id
    where school_class.academic_year='%s' AND school_mark.class_id='%d' AND school_mark.section_id='%d' 
    AND school_mark.staff_id='%d' AND school_class.school_id_id='%d' GROUP BY school_subject.subject_id''' % (academic_year,class_id,section_id,int(staff_id),int(school_id))
    sub = cursor.execute(subject)
    row = cursor.fetchall()

    data = {
    'row':row,
    'student_count':student_count
    }
    return JsonResponse(data)

def single_student_subject_marks_chart(request,pk):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
    	school_id=request.session['schoolname']
    	student_name = StudentDetail.objects.filter(register_number=pk)
    	cursor = connection.cursor()
    	post = ''' SELECT SUM(school_mark.mark),school_mark.subject_id
    	from school_mark INNER JOIN school_subject ON 
    	school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' GROUP BY school_mark.subject_id ''' % (pk)
    	sub = cursor.execute(post)
    	row = cursor.fetchall()
    	lenth = len(row)

    	subject_name = '''SELECT COUNT(school_subject.subject_id),school_subject.subject_name,school_subject.subject_id from school_mark INNER JOIN school_subject ON 
    	school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' 
    	GROUP BY school_subject.subject_name,school_subject.subject_id ''' % (pk)
    	query = cursor.execute(subject_name)
    	sub_name =cursor.fetchall()

    	return render(request,'single_student_subject_marks_chart.html',{'mark':row,'student_name':student_name,'lenth':lenth,'subject_name':sub_name})
    else:
        return redirect("teacher_login")
def student_subject_marks_chart(request,pk):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
        student_name = StudentDetail.objects.filter(register_number=pk)
        cursor = connection.cursor()
        mark = ''' SELECT SUM(school_mark.mark),school_mark.subject_id
        from school_mark INNER JOIN school_subject ON 
        school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' GROUP BY school_mark.subject_id ''' % (pk)
        sub = cursor.execute(mark)
        row = cursor.fetchall()
        lenth = len(row)
        subject_name = '''SELECT COUNT(school_subject.subject_id),school_subject.subject_name,school_subject.subject_id from school_mark INNER JOIN school_subject ON 
        school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' 
        GROUP BY school_subject.subject_name,school_subject.subject_id ''' % (pk)
        query = cursor.execute(subject_name)
        sub_name =cursor.fetchall()
        return render(request,'student_subject_marks_chart.html',{'mark':row,'student_name':student_name,'lenth':lenth,'subject_name':sub_name})
    else:
        return redirect("student_login")
def edit_students_mark(request,pk):
    staff_id=request.session['user_id']
    academic_year = request.session['academic_year']
    cursor = connection.cursor()
    mark = '''SELECT * from school_mark INNER JOIN school_class ON school_mark.class_id=school_class.class_id
    where school_mark.student_reg_no='%s' AND school_mark.staff_id='%d' 
    AND school_class.academic_year='%s' ''' % (pk,int(staff_id),academic_year)
    query = cursor.execute(mark)
    row = cursor.fetchall()
    return render(request,'edit_students_mark.html',{'row':row})

def ajax_subject_count(request):
    student_reg_no = request.GET.get('ids')
    cursor = connection.cursor()
    post = ''' SELECT SUM(school_subject.subject_id),school_subject.subject_name,school_subject.subject_id
    from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' GROUP BY school_subject.subject_id''' % (student_reg_no)
    sub = cursor.execute(post)
    row = cursor.fetchall()

    post2 = ''' SELECT COUNT(school_mark.subject_id) from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.student_reg_no='%s' GROUP BY school_mark.subject_id''' % (student_reg_no)
    sub1 = cursor.execute(post2)
    row1 = cursor.fetchall()

    lenth = len(row)
    data = {
    'lenth':lenth,
    'row':row,
    'row1':row1
    }
    return JsonResponse(data)
def teacher_class_diary(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        school_id=request.session['schoolname']
        academic_year=request.session['academic_year']
        staff_id=request.session['user_id']
        query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year).select_related('class').select_related('section')
        query = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('section_id'), count2= Count('class_id')).order_by('class_id__class_name')
        return render(request,'teacher_class_diary.html',{'staff':query})
    else:
        return redirect("teacher_login")
def manage_exam_mark_chart(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        school_id=request.session['schoolname']
        academic_year=request.session['academic_year']
        staff_id=request.session['user_id']
        query = AssignSubjectTeacher.objects.filter(staff_id_id=staff_id,academic_year=academic_year)
        query = query.values('class_id', 'section_id','class_id__class_name','section_id__section_name').annotate(count = Count('class_id')).order_by('class_id__class_name')
        return render(request,'manage_exam_mark_chart.html',{'staff':query})
    else:
        return redirect("teacher_login")
def particular_exam_mark_chart(request,pk,cls_id,sec_id):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        cursor = connection.cursor()
        post = ''' SELECT school_subject.subject_id, school_subject.subject_name  from school_subject INNER JOIN school_mark  
        ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
        where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
        GROUP BY school_subject.subject_id order by school_subject.subject_id'''  % (cls_id,sec_id,pk)
        query = cursor.execute(post)
        row = cursor.fetchall()

        post1 = ''' SELECT school_mark.mark, school_mark.subject_id,school_mark.exam_id from school_subject INNER JOIN school_mark  
        ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
        where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
        order by school_exam.exam_id,school_subject.subject_id'''  % (cls_id,sec_id,pk)
        query1 = cursor.execute(post1)
        row1 = cursor.fetchall()

        post2 = ''' SELECT COUNT(school_exam.exam_id),school_exam.exams, SUM(school_mark.mark) from school_mark INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
        where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
        Group By school_exam.exam_id order by school_exam.exam_id '''  % (cls_id,sec_id,pk)
        query2 = cursor.execute(post2)
        row2 = cursor.fetchall()

        post5 = ''' SELECT COUNT(school_mark.exam_id) from school_mark 
        where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
        Group By school_mark.exam_id '''  % (cls_id,sec_id,pk)
        query5 = cursor.execute(post5)
        row5 = cursor.fetchall()
        count3 = len(row5)

        exam_id = Exam.objects.filter(class_id=cls_id,section_id=sec_id).order_by('exam_id')
        exam = Mark.objects.filter(student_reg_no=pk)
        cnt = exam.values('exam_id','student_reg_no').annotate(count=Count('exam_id'),count1=Count('student_reg_no'))

        return render(request,'particular_exam_mark_chart.html',{'row':row,'length':cnt,'exam_id':exam_id,'row1':row1,'row2':row2,'pk':pk,
        'len_exam':count3})
    else:
        return redirect("teacher_login")
def get_exam_name(request):
    cls_id = request.POST.get('cls_id')
    sec_id = request.POST.get('sec_id')
    pk = request.POST.get('student_id')
    class_id = int(cls_id)
    section_id = int(sec_id)
    cursor = connection.cursor()
    post = ''' SELECT school_mark.mark, school_mark.subject_id,school_mark.exam_id, school_exam.exams from school_subject INNER JOIN school_mark  
    ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
    where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
    order by school_exam.exam_id,school_subject.subject_id'''  % (class_id,section_id,pk)
    query = cursor.execute(post)
    row = cursor.fetchall()

    data = {
    'row':row
    }
    return JsonResponse(data)
def Student_particular_exam_mark_chart(request,pk,cls_id,sec_id):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
        cursor = connection.cursor()
        post = ''' SELECT school_subject.subject_id, school_subject.subject_name  from school_subject INNER JOIN school_mark  
        ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
        where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
        GROUP BY school_subject.subject_id order by school_subject.subject_id'''  % (cls_id,sec_id,pk)
        query = cursor.execute(post)
        row = cursor.fetchall()

        post1 = ''' SELECT school_mark.mark, school_mark.subject_id,school_mark.exam_id from school_subject INNER JOIN school_mark  
        ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
        where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
        order by school_exam.exam_id,school_subject.subject_id'''  % (cls_id,sec_id,pk)
        query1 = cursor.execute(post1)
        row1 = cursor.fetchall()

        post2 = ''' SELECT COUNT(school_exam.exam_id),school_exam.exams, SUM(school_mark.mark) from school_mark INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
        where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
        Group By school_exam.exam_id order by school_exam.exam_id '''  % (cls_id,sec_id,pk)
        query2 = cursor.execute(post2)
        row2 = cursor.fetchall()

        post5 = ''' SELECT COUNT(school_mark.exam_id) from school_mark 
        where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
        Group By school_mark.exam_id '''  % (cls_id,sec_id,pk)
        query5 = cursor.execute(post5)
        row5 = cursor.fetchall()
        count3 = len(row5)

        exam_id = Exam.objects.filter(class_id=cls_id,section_id=sec_id).order_by('exam_id')
        exam = Mark.objects.filter(student_reg_no=pk)
        cnt = exam.values('exam_id','student_reg_no').annotate(count=Count('exam_id'),count1=Count('student_reg_no'))

        return render(request,'Student_particular_exam_mark_chart.html',{'row':row,'length':cnt,'exam_id':exam_id,'row1':row1,'row2':row2,'pk':pk,
        'len_exam':count3})
    else:
        return redirect("student_login")

def select_exam(request):
    cls_id = request.POST.get('class_id')
    class_id = int(cls_id)
    sec_id = request.POST.get('section_id')
    section_id = int(sec_id)
    staff_id = request.session['user_id']
    cursor = connection.cursor()
    sql= ''' SELECT * from school_exam where school_exam.class_id_id='%d' AND school_exam.section_id_id='%d'
    ''' % (class_id,section_id)
    res = cursor.execute(sql)
    exam_name = cursor.fetchall()
    post = ''' SELECT COUNT(school_mark.exam_id) ,COUNT(school_mark.subject_id),SUM(school_mark.mark),COUNT(school_mark.student_reg_no) from school_mark  where school_mark.class_id='%d' AND school_mark.section_id='%d'
    AND school_mark.staff_id='%d' GROUP BY school_mark.subject_id''' % (class_id,section_id,staff_id)
    query = cursor.execute(post)
    row = cursor.fetchall()
    data = {
    'row':row,
    'exam_name':exam_name
    }
    return JsonResponse(data)

def select_school_exam(request):
    cls_id = request.POST.get('class_id')
    class_id = int(cls_id)
    sec_id = request.POST.get('section_id')
    section_id = int(sec_id)
    cursor = connection.cursor()
    sql= ''' SELECT  app_exam.class_id_id,app_exam.section_id_id,app_exam.exam_id,app_exam.exams from app_exam where app_exam.class_id_id='%d' AND app_exam.section_id_id='%d'
    ''' % (class_id,section_id)
    res = cursor.execute(sql)
    exam_name = cursor.fetchall()
    data = {
    'exam_name':exam_name
    }
    return JsonResponse(data)

def select_school_details(request):
    school_id = request.session['schoolname']
    cursor = connection.cursor()
    sql= ''' SELECT * from school_schooldetail where school_schooldetail.id='%d' ''' % (school_id)
    res = cursor.execute(sql)
    school_details = cursor.fetchall()
    data ={
        'school_details':school_details
    }
    return JsonResponse(data)

def search_student_marks_ajax(request):
    cls_id = request.POST.get('class_id')
    class_id = int(cls_id)
    sec_id = request.POST.get('section_id')
    section_id = int(sec_id)
    ex_id = request.POST.get('exam_id')
    exam_id = int(ex_id)
    student_reg_no = request.POST.get('student_reg_no')
    cursor = connection.cursor()
    stud = '''SELECT * from app_studentdetail'''
    result = cursor.execute(stud)
    student_id = cursor.fetchall()

    if  student_reg_no == "":
        sql= '''SELECT d.student_name,m.reg_no,m.student_id_id, s.class_id_id, s.section_id_id,m.exam_id_id from app_studentdetail as d INNER JOIN app_studentsection as s ON d.id=s.student_id_id
		INNER JOIN app_mark as m ON s.student_id_id=m.student_id_id where s.class_id_id='%d' AND s.section_id_id='%d' GROUP BY d.id ''' % (class_id,section_id)
        res = cursor.execute(sql)
        mark_details = cursor.fetchall()
        data ={
            'mark_details':mark_details,
            'student_id': student_id,
            'class_id':class_id,
            'section_id':section_id,
            'exam_id':exam_id
        }
    else:
        sql= ''' SELECT d.student_name,m.reg_no,m.student_id_id from app_studentdetail as d INNER JOIN app_studentsection as s ON d.id=s.student_id_id
		INNER JOIN app_mark as m ON s.student_id_id=m.student_id_id where s.class_id_id='%d' AND s.section_id_id='%d'
		AND m.reg_no='%s' ''' % (class_id,section_id,student_reg_no)
        res = cursor.execute(sql)
        mark_details = cursor.fetchall()
        data ={
            'mark_details':mark_details,
            'class_id':class_id,
            'section_id':section_id,
            'exam_id':exam_id
        }
    return JsonResponse(data)

def landing_page(request):
    return render(request,'home.html',{})
from django.db.models.query import QuerySet
def student_class_mark(request,class_id,section_id,exam_id):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        if request.method == "POST":
            ex_id = request.POST.get('ex_id')
            stud_id = request.POST.get('stud_id')
            exam_id = int(ex_id)
            student_id = int(stud_id)
            mark_delete = Mark.objects.filter(exam_id=exam_id,student_reg_no=student_id)
            mark_delete.delete()
        school_id = request.session['schoolname']
        staff_id = request.session['user_id']
        academic_year = request.session['academic_year']
        cursor = connection.cursor()

        stud_ids = ''' SELECT school_mark.student_reg_no, COUNT(school_mark.student_reg_no) from school_mark INNER JOIN
        school_studentdetail ON school_mark.student_reg_no=school_studentdetail.register_number
        WHERE school_mark.class_id='%d' AND school_mark.section_id='%d'
        AND school_mark.exam_id='%d' GROUP BY school_mark.student_reg_no order by school_mark.student_reg_no''' % (class_id,section_id,exam_id)
        exams_id = cursor.execute(stud_ids)
        exams_name = cursor.fetchall()

        a  = Exam.objects.filter(exam_id=exam_id)
        subject_name = AssignSubjectTeacher.objects.filter(class_id=class_id,section_id=section_id,staff_id=staff_id).order_by('subject_id')
        results = Mark.objects.filter(exam_id=exam_id,class_id=class_id,section_id=section_id).select_related('exam').select_related('studentsection')
        cnt = results.values('student_reg_no','exam_id','mark').annotate(count=Count('student_reg_no'),count1=Count('exam_id'))
        page = request.GET.get('page', 1)
        page_count = pagination_cnt(request)
        paginator = Paginator(exams_name,page_count)
        try:
            users = paginator.page(page)
        except PageNotAnInteger:
            users = paginator.page(1)
        except EmptyPage:
            users = paginator.page(paginator.num_pages)

        return render(request,'student_class_mark.html',{'exams_name':users,'a':a,'results':cnt,'subject_name':subject_name,'class_id':class_id,
        'section_id':section_id,'exam_id':exam_id})
    else:
        return redirect("teacher_login")
def mark_already_exist(request):
	cls_id = request.POST.getlist('class_id[]')
	sec_id = request.POST.getlist('section_id[]')
	staff_id = request.POST.getlist('staff_id[]')
	student_id = request.POST.getlist('stud_reg_no[]')
	subject_id = request.POST.getlist('subject_id[]')
	ex_id = request.POST.get('exams')
	exam_id = int(ex_id)
	length = len(subject_id)
	cursor = connection.cursor()

	for i in range(0,length):
		if subject_id[i]!= "":
			post = '''SELECT * from school_mark where school_mark.class_id='%d' AND school_mark.section_id='%d' AND 
			school_mark.subject_id='%d' AND school_mark.exam_id='%d' AND school_mark.staff_id='%d' AND 
			school_mark.student_reg_no='%s' ''' % (int(cls_id[i]),int(sec_id[i]),int(subject_id[i]),exam_id,int(staff_id[i]),student_id[i])
			query = cursor.execute(post)
			row = cursor.fetchall()
			data = {
			'msg':"Mark Already Exist",
			'row':row,
			'length':student_id
			}
			return JsonResponse(data)
def choose_exam(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
    	if request.GET.get('cls_id') and request.GET.get('sec_id'):
    		cls_id = request.GET.get('cls_id')
    		sec_id = request.GET.get('sec_id')
    		class_id = int(cls_id)
    		section_id = int(sec_id)
    		school_id = request.session['schoolname']
    		exams = Exam.objects.filter(class_id_id=class_id,section_id_id=section_id,school_id_id=school_id)
    		return render(request,'choose_exam.html',{'exams':exams})
    	else:
    		return redirect("manage_students_marks")
    else:
        return redirect("teacher_login")
def exam_mark_chart(request,cls_id,sec_id,pk):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        stud_id = pk
        school_id = request.session['schoolname']
        exams = Exam.objects.filter(class_id_id=cls_id,section_id_id=sec_id,school_id_id=school_id)
        return render(request,'exam_mark_chart.html',{'exams':exams,'pk':pk})
    else:
        return redirect("teacher_login")
def single_student_exam_mark_chart(request,cls_id,sec_id,pk):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
        stud_id = pk
        school_id = request.session['schoolname']
        exams = Exam.objects.filter(class_id_id=cls_id,section_id_id=sec_id,school_id_id=school_id)
        return render(request,'single_student_exam_mark_chart.html',{'exams':exams,'pk':pk})
    else:
        return redirect("student_login")
def exam_chart(request,cls_id,sec_id,exam_id,stud_id):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        exam = Exam.objects.filter(exam_id=exam_id)
        student_name = StudentDetail.objects.filter(register_number=stud_id)
        return render(request,'exam_chart.html',{'exam':exam,'student_name':student_name,'cls_id':cls_id,
        'sec_id':sec_id,'stud_id':stud_id,'exam_id':exam_id})
    else:
        return redirect("teacher_login")
def particular_exam_chart(request,cls_id,sec_id,exam_id,stud_id):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
        exam = Exam.objects.filter(exam_id=exam_id)
        student_name = StudentDetail.objects.filter(register_number=stud_id)
        return render(request,'particular_exam_chart.html',{'exam':exam,'student_name':student_name,'cls_id':cls_id,
        'sec_id':sec_id,'stud_id':stud_id,'exam_id':exam_id})
    else:
        return redirect("student_login")
def sub_ajax(request):
    class_id = request.POST.get('class_id')
    cls_id = int(class_id)
    section_id = request.POST.get('section_id')
    sec_id = int(section_id)
    exam_id = request.POST.get('exam_id')
    ex_id = int(exam_id)
    stud_id = request.POST.get('std_id')
    cursor = connection.cursor()
    sql= ''' SELECT school_subject.subject_name, school_mark.mark from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.class_id='%d' AND school_mark.section_id='%d' 
    AND school_mark.exam_id='%d' AND school_mark.student_reg_no='%s' ''' % (cls_id,sec_id,ex_id,stud_id)
    res = cursor.execute(sql)
    mark_details = cursor.fetchall()

    data = {
    'row':mark_details
    }
    return JsonResponse(data)
def select_exams(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
    	if request.GET.get('cls_id') and request.GET.get('sec_id'):
    		cls_id = request.GET.get('cls_id')
    		sec_id = request.GET.get('sec_id')
    		class_id = int(cls_id)
    		section_id = int(sec_id)
    		school_id = request.session['schoolname']
    		exams = Exam.objects.filter(class_id_id=class_id,section_id_id=section_id,school_id_id=school_id)
    		return render(request,'select_exams.html',{'exams':exams})
    	else:
    		return redirect("manage_students_marks")
    else:
        return redirect("teacher_login")
def select_exam_excel(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        if request.GET.get('cls_id') and request.GET.get('sec_id'):
            cls_id = request.GET.get('cls_id')
            sec_id = request.GET.get('sec_id')
            class_id = int(cls_id)
            section_id = int(sec_id)
            school_id = request.session['schoolname']
            exams = Exam.objects.filter(class_id_id=class_id,section_id_id=section_id,school_id_id=school_id)
            return render(request,'select_exam_excel.html',{'exams':exams})
        else:
            return redirect("manage_students_marks")
    else:
        return redirect("teacher_login")
def mark_ajax(request):
    class_id = request.POST.get('class_id')
    section_id = request.POST.get('section_id')
    exam_id = request.POST.get('exam_id')
    teacher_id = request.session['user_id']
    cursor = connection.cursor()
    stud_ids = ''' SELECT school_mark.student_reg_no,school_mark.mark,school_mark.subject_id, school_studentdetail.student_name from school_mark  INNER JOIN school_studentdetail ON 
    school_mark.student_reg_no=school_studentdetail.register_number
    WHERE school_mark.class_id='%d' AND school_mark.section_id='%d'
    AND school_mark.exam_id='%d' AND school_mark.staff_id='%d' order by school_mark.subject_id''' % (int(class_id),int(section_id),int(exam_id),int(teacher_id))
    exams_id = cursor.execute(stud_ids)
    exams_name = cursor.fetchall()

    post = ''' SELECT school_studentdetail.student_name,school_studentdetail.register_number from school_studentdetail
    INNER JOIN school_studentsection ON school_studentdetail.id=school_studentsection.student_id_id 
    where school_studentsection.class_id_id='%d' AND school_studentsection.section_id_id='%d'  
    order by school_studentdetail.id''' % (int(class_id),int(section_id))
    student_id = cursor.execute(post)
    student_name = cursor.fetchall()


    data = {
    'row':exams_name,
    'student_name':student_name
    }
    return JsonResponse(data)
def teacher_mark_ajax(request):
    class_id = request.POST.get('class_id')
    section_id = request.POST.get('section_id')
    exam_id = request.POST.get('exam_id')
    cursor = connection.cursor()
    stud_ids = ''' SELECT school_mark.student_reg_no,school_mark.mark,school_mark.subject_id, school_studentdetail.student_name from school_mark  INNER JOIN school_studentdetail ON 
    school_mark.student_reg_no=school_studentdetail.register_number
    WHERE school_mark.class_id='%d' AND school_mark.section_id='%d'
    AND school_mark.exam_id='%d' order by school_mark.subject_id''' % (int(class_id),int(section_id),int(exam_id))
    exams_id = cursor.execute(stud_ids)
    exams_name = cursor.fetchall()


    data = {
    'row':exams_name
    }
    return JsonResponse(data)
def delete_mark(request):
    stud_id = request.POST.get('stud_id')
    exam_id = request.POST.get('ex_id')
    cursor = connection.cursor()
    stud_ids = ''' DELETE from school_mark 
    WHERE school_mark.exam_id='%d' AND school_mark.student_reg_no='%s' ''' % (int(exam_id),stud_id)
    exams_id = cursor.execute(stud_ids)

    data = {
    'row':'deleted'
    }
    return JsonResponse(data)
def view_student_mark(request):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
        stud_id = request.session['student_id']
        student = StudentSection.objects.filter(student_id=stud_id)
        return render(request,'view_student_mark.html',{'stud_id':student})
    else:
        return redirect("student_login")
def exam_list(request):
    cls_id = request.POST.get('cls_id')
    sec_id = request.POST.get('sec_id')
    cursor = connection.cursor()
    stud_ids = ''' SELECT * from school_exam 
    WHERE school_exam.class_id_id='%d' AND school_exam.section_id_id='%d'
    order by school_exam.exam_id''' % (int(cls_id),int(sec_id))
    exams_id = cursor.execute(stud_ids)
    exams_name = cursor.fetchall()
    data = {
    'row':exams_name
    }
    return JsonResponse(data)
def view_single_student_particular_exam_mark(request):
    if request.session.has_key('student') and request.session.has_key('academic_year'):
        tot = 0
        Register_number = request.session['register_number']
        exam_id =""
        if request.GET.get('id'):
            exam_id = request.GET.get('id')
        cursor = connection.cursor()
        Mark = ''' SELECT school_studentdetail.student_name, school_mark.*,school_mark.subject_id, school_subject.subject_name , school_exam.exams from school_mark
        LEFT JOIN school_subject ON school_mark.subject_id=school_subject.subject_id
        LEFT JOIN school_exam ON school_mark.exam_id=school_exam.exam_id
        LEFT JOIN school_studentdetail ON school_studentdetail.register_number=school_mark.student_reg_no
        WHERE school_mark.exam_id='%d' AND school_mark.student_reg_no='%s'
        ''' % (int(exam_id),Register_number)
        Mark_id = cursor.execute(Mark)
        Marks = cursor.fetchall()
        length = len(Marks)
        for i in Marks:
            tot+=i[7]
        average = tot/length
        if(average>=91 and average<=100):
            grade = "A"
        elif(average>=81 and average<=90.9):
            grade = "B"
        elif(average>=71 and average<=80.9):
            grade = "C"
        elif(average>=61 and average<=70.9):
            grade = "D"
        elif(average>=51 and average<=60.9):
            grade = "E"
        elif(average>=41 and average<=50.9):
            grade = "O"
        elif(average>=0 and average<=40.9):
            grade = "F"

        if grade == "F":
            result = "Fail"
        else:
            result = "Pass"
        percent = format(average, '.1f')
        return render(request,'view_single_student_particular_exam_mark.html',{'Register_number':Register_number,'exam':exam_id,'Marks':Marks,'tot':tot,'length':length,'percentage':percent,'grade':grade,'result':result})
    else:
        return redirect("student_login")

def ajax_search_particular_student_mark(request):
    if request.POST.get('reg_no') != "" and request.POST.get('exam_id_id') !="":
        exam_id = request.POST.get('exam_id_id')
        reg_no = request.POST.get('reg_no')
        teacher_id = request.session['user_id']
        cursor = connection.cursor()
        Mark = ''' SELECT school_studentdetail.register_number, school_studentdetail.student_name, school_mark.mark,school_mark.subject_id from school_mark
        LEFT JOIN school_subject ON school_mark.subject_id=school_subject.subject_id
        LEFT JOIN school_exam ON school_mark.exam_id=school_exam.exam_id
        LEFT JOIN school_studentdetail ON school_studentdetail.register_number=school_mark.student_reg_no
        WHERE school_mark.exam_id='%d' AND school_mark.student_reg_no='%s' AND school_mark.staff_id='%d'
        ''' % (int(exam_id),reg_no,int(teacher_id))
        Mark_id = cursor.execute(Mark)
        Marks = cursor.fetchall()
        data = {
            'Marks':Marks
        }
    return JsonResponse(data)
def ajax_search_all_student_mark(request):
    if request.POST.get('reg_no') != "" and request.POST.get('exam_id_id') !="":
        tot = 0
        average = 0
        exam_id = request.POST.get('exam_id_id')
        reg_no = request.POST.get('reg_no')
        teacher_id = request.session['user_id']
        cursor = connection.cursor()
        Mark = ''' SELECT school_studentdetail.register_number, school_studentdetail.student_name, school_mark.mark,school_mark.subject_id from school_mark
        LEFT JOIN school_subject ON school_mark.subject_id=school_subject.subject_id
        LEFT JOIN school_exam ON school_mark.exam_id=school_exam.exam_id
        LEFT JOIN school_studentdetail ON school_studentdetail.register_number=school_mark.student_reg_no
        WHERE school_mark.exam_id='%d' AND school_mark.student_reg_no='%s' 
        ''' % (int(exam_id),reg_no)
        Mark_id = cursor.execute(Mark)
        Marks = cursor.fetchall()
        length = len(Marks)
        for i in Marks:
            tot+=i[2]
        if tot > 0:
            average = tot/length
        if(average>=91 and average<=100):
            grade = "A"
        elif(average>=81 and average<=90.9):
            grade = "B"
        elif(average>=71 and average<=80.9):
            grade = "C"
        elif(average>=61 and average<=70.9):
            grade = "D"
        elif(average>=51 and average<=60.9):
            grade = "E"
        elif(average>=41 and average<=50.9):
            grade = "O"
        elif(average>=0 and average<=40.9):
            grade = "F"

        if grade == "F":
            result = "Fail"
        else:
            result = "Pass"
        percent = format(average, '.1f')
        data = {
            'Marks':Marks,
            'average':average,
            'grade':grade,
            'tot':tot
        }
    return JsonResponse(data)
def test_teacher(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        teacher = StaffDetail.objects.filter(school_id=school_id)
        class_id = Class.objects.filter(school_id=school_id,academic_year=academic_year)
        if request.method == "POST":
            form = AssignTeacherForm(request.POST)
            if form.is_valid():
                post = form.save(commit=False)
                post.save()
                return redirect("assign_teachers_view")  
        else:
            form = AssignTeacherForm()
        return render(request, 'test_teacher.html', {'form':form,'teacher':teacher,'class_id':class_id})
    else:
        return redirect("adminlogin")
def student_mark_yearly_chart(request):
	if request.session.has_key('teacher') and request.session.has_key('academic_year'):
		if request.GET.get('exam_id') and request.GET.get('stud_id'):
			ex_id = request.GET.get('exam_id')
			student_id = request.GET.get('stud_id')
			exam = Exam.objects.filter(exam_id=int(ex_id))
			student_name = StudentDetail.objects.filter(register_number=student_id)
			return render(request, 'student_mark_yearly_chart.html', {'exam':exam,'student_name':student_name})
		else:
			return redirect("search_student_year")
	else:
		return redirect("teacher_login")

def sub_ajax_search_mark(request):
    exam_id = request.POST.get('exam_id')
    ex_id = int(exam_id)
    stud_id = request.POST.get('std_id')
    cursor = connection.cursor()
    sql= ''' SELECT school_subject.subject_name, school_mark.mark from school_mark INNER JOIN school_subject ON 
    school_mark.subject_id=school_subject.subject_id where school_mark.exam_id='%d' AND school_mark.student_reg_no='%s' ''' % (ex_id,stud_id)
    res = cursor.execute(sql)
    mark_details = cursor.fetchall()

    data = {
    'row':mark_details
    }
    return JsonResponse(data)

def student_mark_overall_yearly_chart(request):
	if request.session.has_key('teacher') and request.session.has_key('academic_year'):
		if request.GET.get('cls_id') and request.GET.get('sec_id') and request.GET.get('stud_id'):
			cls_id = request.GET.get('sec_id')
			sec_id = request.GET.get('sec_id')
			stud_id = request.GET.get('stud_id')
			cursor = connection.cursor()
			post = ''' SELECT school_subject.subject_id, school_subject.subject_name  from school_subject INNER JOIN school_mark  
			ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
			where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
			GROUP BY school_subject.subject_id order by school_subject.subject_id'''  % (int(cls_id),int(sec_id),stud_id)
			query = cursor.execute(post)
			row = cursor.fetchall()

			post1 = ''' SELECT school_mark.mark, school_mark.subject_id,school_mark.exam_id from school_subject INNER JOIN school_mark  
			ON school_subject.subject_id=school_mark.subject_id INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
			where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
			order by school_exam.exam_id,school_subject.subject_id'''  % (int(cls_id),int(sec_id),stud_id)
			query1 = cursor.execute(post1)
			row1 = cursor.fetchall()

			post2 = ''' SELECT COUNT(school_exam.exam_id),school_exam.exams, SUM(school_mark.mark) from school_mark INNER JOIN school_exam ON school_mark.exam_id = school_exam.exam_id 
			where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
			Group By school_exam.exam_id order by school_exam.exam_id '''  % (int(cls_id),int(sec_id),stud_id)
			query2 = cursor.execute(post2)
			row2 = cursor.fetchall()

			post5 = ''' SELECT COUNT(school_mark.exam_id) from school_mark 
			where school_mark.class_id='%d' AND school_mark.section_id='%d' AND school_mark.student_reg_no='%s'
			Group By school_mark.exam_id '''  % (int(cls_id),int(sec_id),stud_id)
			query5 = cursor.execute(post5)
			row5 = cursor.fetchall()
			count3 = len(row5)

			exam_id = Exam.objects.filter(class_id=int(cls_id),section_id=int(sec_id)).order_by('exam_id')
			exam = Mark.objects.filter(student_reg_no=stud_id)
			cnt = exam.values('exam_id','student_reg_no').annotate(count=Count('exam_id'),count1=Count('student_reg_no'))

			return render(request,'student_mark_overall_yearly_chart.html',{'row':row,'length':cnt,'exam_id':exam_id,'row1':row1,'row2':row2,'pk':stud_id,
			'len_exam':count3})
		else:
			return redirect("search_student_year")
	else:
		return redirect("teacher_login")
def manage_single_student_particular_exam_mark(request):
	if request.session.has_key('teacher') and request.session.has_key('academic_year'):
		tot = 0
		exam_id =""
		if request.GET.get('id') and request.GET.get('exam_id'):
			student_id = request.GET.get('id')
			exam_id = request.GET.get('exam_id')
			cursor = connection.cursor()
			Mark = ''' SELECT school_subject.subject_name,school_mark.mark, school_exam.exams,school_studentdetail.student_name from school_mark
			LEFT JOIN school_subject ON school_mark.subject_id=school_subject.subject_id
			LEFT JOIN school_exam ON school_mark.exam_id=school_exam.exam_id
			LEFT JOIN school_studentdetail ON school_studentdetail.register_number=school_mark.student_reg_no
			WHERE school_mark.exam_id='%d' AND school_mark.student_reg_no='%s'
			''' % (int(exam_id),student_id)
			Mark_id = cursor.execute(Mark)
			Marks = cursor.fetchall()
			length = len(Marks)
			for i in Marks:
				tot+=i[1]
			average = tot/length
			if(average>=91 and average<=100):
				grade = "A"
			elif(average>=81 and average<=90.9):
				grade = "B"
			elif(average>=71 and average<=80.9):
				grade = "C"
			elif(average>=61 and average<=70.9):
				grade = "D"
			elif(average>=51 and average<=60.9):
				grade = "E"
			elif(average>=41 and average<=50.9):
				grade = "O"
			elif(average>=0 and average<=40.9):
				grade = "F"

			if grade == "F":
				result = "Fail"
			else:
				result = "Pass"
			percent = format(average, '.1f')
			return render(request,'manage_single_student_particular_exam_mark.html',{'exam':exam_id,'Marks':Marks,'tot':tot,'length':length,'percentage':percent,'grade':grade,'result':result})
		else:
			return redirect("search_student_year")
	else:
		return redirect("teacher_login")
def add_parents(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        stud_detail = StudentSection.objects.filter(student_id__in=StudentDetail.objects.filter(school_id=school_id))
        row = ParentsDetail.objects.filter(school_id=school_id)
        if request.method == "POST":
            uname = request.POST.get('uname')
            psw = request.POST.get('psw')
            stud_id = request.POST.get('student_id')
            ids =  StudentSection.objects.get(student_id=int(stud_id))
            reg = request.POST.get('reg')
            email = request.POST.get('email')
            sch_id = SchoolDetail.objects.get(id=int(school_id))
            sub = ParentsDetail.objects.create(student_id=ids,username=uname,password=psw,register_number=reg,
            email=email,school_id=sch_id)
            if sub:
                messages.success(request,"Parent's Detail Added Successfully")
                return redirect("add_parents")
        else:
            form = SubjectForm()
        return render(request, 'add_parents.html', {'row':row,'stud_detail':stud_detail})
    else:
        return redirect("adminlogin")
def add_events(request):
    if request.session.has_key('username') and request.session.has_key('academic_year'):
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        row = EventDetail.objects.filter(school_id=school_id)
        if request.method == "POST":
            name = request.POST.get('name')
            date = request.POST.get('date')
            time = request.POST.get('time')
            desc = request.POST.get('desc')
            addr = request.POST.get('addr')
            sch_id = SchoolDetail.objects.get(id=int(school_id))
            sub = EventDetail.objects.create(event_name=name,date=date,time=time,desc=desc,
            address=addr,school_id=sch_id)
            if sub:
                messages.success(request,"Event Detail Added Successfully")
                return redirect("add_events")
        else:
            form = SubjectForm()
        return render(request, 'add_events.html', {'row':row})
    else:
        return redirect("adminlogin")
def delete_event(request,pk):
    EventDetail.objects.filter(id=int(pk)).delete()
    return redirect('add_events')
def delete_parent(request,pk):
    ParentsDetail.objects.filter(id=int(pk)).delete()
    return redirect('add_parents')
def send_meeting(request):
    if request.session.has_key('teacher') and request.session.has_key('academic_year'):
        school_id = request.session['schoolname']
        academic_year = request.session['academic_year']
        teacher_id = request.session['user_id']
        sid =  SchoolDetail.objects.get(id=int(school_id))
        result = Meeting.objects.filter(staff_id=teacher_id)
        a = StaffDetail.objects.get(staff_id=int(teacher_id))
        row = AssignSubjectTeacher.objects.filter(staff_id=int(teacher_id))
        if request.method == "POST":
            cls_id = request.POST.get('cls_id')
            sec_id = request.POST.get('sec_id')
            date = request.POST.get('date')
            time = request.POST.get('time')
            msg = request.POST.get('msg')
            class_id = Class.objects.get(class_id=int(cls_id))
            section_id = Section.objects.get(section_id=int(sec_id))
            sub = Meeting.objects.create(date=date,time=time,msg=msg,
            class_id=class_id,section_id=section_id,staff_id=a,school_id=sid)
            if sub:
                messages.success(request,"Message Sent to Parents Successfully")
                return redirect("send_meeting")
        else:
            form = SubjectForm()
        return render(request, 'send_meeting.html', {'row':row,'result':result,'staff_id':a})
    else:
        return redirect("teacher_login")
def delete_meeting(request,pk):
    Meeting.objects.filter(id=int(pk)).delete()
    return redirect('send_meeting')
def parent_login(request):
    if request.session.has_key('parent') and request.session.has_key('academic_year'):
        return redirect("parent_dashboard")
    else:
        if request.method == 'POST':
            username = request.POST['username']
            password =  request.POST['password']
            post = ParentsDetail.objects.filter(username=username,password=password)
            if post:
                username = request.POST['username']
                request.session['parent'] = username
                a = request.session['parent']
                user_id = ParentsDetail.objects.only('id').get(username=a).id
                request.session['parent_id']=user_id
                sess = ParentsDetail.objects.only('school_id_id').get(id=user_id).school_id_id
                request.session['schoolname']=sess
                return redirect("parent_dashboard")
            else:
                messages.success(request, 'Invalid Username or Password')
    return render(request, 'parent_login.html', {})
def parent_dashboard(request):
    if request.session.has_key('parent'):
        return render(request, 'parent_dashboard.html',{})
    else:
        return render(request, 'parent_login.html',{})
def parent_logout(request):
    try:
        del request.session['parent']
        del request.session['academic_year']
    except:
     pass
    return render(request, 'parent_login.html', {})
def view_student_mark_parent(request):
    if request.session.has_key('parent') and request.session.has_key('academic_year'):
        stud_id = request.session['parent_id']
        student = ParentsDetail.objects.filter(id=int(stud_id))
        return render(request,'view_student_mark_parent.html',{'stud_id':student})
    else:
        return redirect("parent_login")
def metting_detail(request,cid,sid):
    if request.session.has_key('parent') and request.session.has_key('academic_year'):
        stud_id = request.session['parent_id']
        student = Meeting.objects.filter(class_id=int(cid),section_id=int(sid))
        return render(request,'metting_detail.html',{'stud_id':student})
    else:
        return redirect("parent_login")
def event_details_teacher(request):
    school_id = request.session['schoolname']
    student = EventDetail.objects.filter(school_id=int(school_id))
    return render(request,'event_details_teacher.html',{'stud_id':student})
def event_details_student(request):
    school_id = request.session['schoolname']
    student = EventDetail.objects.filter(school_id=int(school_id))
    return render(request,'event_details_student.html',{'stud_id':student})
def event_details_parent(request):
    school_id = request.session['schoolname']
    student = EventDetail.objects.filter(school_id=int(school_id))
    return render(request,'event_details_parent.html',{'stud_id':student})
#----------------------------------------------------------------------------------------
def admin_class(request):
	if request.session.has_key('username') or request.session.has_key('teacher'):
		school_id = request.session['schoolname']
		academic_year=request.session['academic_year']
		detail = Section.objects.filter(class_id__in=Class.objects.filter(academic_year=academic_year),school_id=int(school_id))
		return render(request, 'admin_class.html', {'detail':detail})
	else:
		return redirect('adminlogin')
def view_exam(request,cls_id,sec_id):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		detail = Exam.objects.filter(class_id=cls_id,section_id=sec_id)
		return render(request, 'view_exam.html', {'detail':detail})
	else:
		return redirect('adminlogin')
def view_all_students(request,cls_id,sec_id,pk):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		detail = StudentSection.objects.filter(class_id=cls_id,section_id=sec_id)
		return render(request, 'view_all_students.html', {'detail':detail,'exam_id':pk})
	else:
		return redirect('adminlogin')
def add_subject_mark(request,cls_id,sec_id,pk):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		exams = request.GET.get('exam_id')
		detail = Subject.objects.filter(class_id=cls_id,section_id=sec_id,exam_id=exams)
		if request.method == 'POST':
			subject_id = request.POST.getlist('subject_id[]')
			mark = request.POST.getlist('mark_id[]')
			student_id = request.POST.get('student_id')
			exam_id = request.POST.get('exam_id')
			sid = StudentDetail.objects.get(id=int(student_id))
			ex = Exam.objects.get(exam_id=int(exam_id))
			exam_type = request.POST.get('exam_type')
			reg = request.POST.get('reg')
			length=len(mark)
			for row in range(0,length):
				if mark[row]=="":
					mark[row]=0;
				crt=Mark.objects.create(student_id=sid,class_id=cls_id,section_id=sec_id,subject_id=subject_id[row],
				exam_id=ex,mark=int(mark[row]),mark_type=exam_type,reg_no=reg) 
				if crt:
					messages.success(request,"Added Successfully")
		return render(request, 'add_subject_mark.html', {'detail':detail,'student_id':pk})
	else:
		return redirect('adminlogin')
def view_subject_mark(request,cls_id,sec_id,pk):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		exam_id = request.GET.get('exam_id')
		cursor = connection.cursor()
		Mark = ''' SELECT s.subject_name,m.mark,m.mark_type from app_mark as m INNER JOIN app_subject as s ON m.subject_id=s.subject_id where
		m.class_id='%d' AND m.section_id='%d' AND m.student_id_id='%d' AND m.exam_id_id='%d' AND m.mark_type='%s' ''' % (cls_id,sec_id,pk,int(exam_id),'Internal')
		Mark_id = cursor.execute(Mark)
		marks = cursor.fetchall()
		post = ''' SELECT AVG(m.mark) from app_mark as m INNER JOIN app_subject as s ON m.subject_id=s.subject_id where
		m.class_id='%d' AND m.section_id='%d' AND m.student_id_id='%d' AND m.exam_id_id='%d' AND m.mark_type='%s' ''' % (cls_id,sec_id,pk,int(exam_id),'Internal')
		sql= cursor.execute(post)
		row = cursor.fetchall()
		exam_name = Exam.objects.filter(exam_id=int(exam_id))
		return render(request, 'view_subject_mark.html', {'marks':marks,'exam_name':exam_name,'row':row})
	else:
		return redirect('adminlogin')
def student_list(request,cls_id,sec_id,pk):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		if request.method == 'POST':
			total_days = request.POST.get('total_days')
			exam_id = request.POST.get('exam_id')
			sid = StudentDetail.objects.get(id=pk)
			ex = Exam.objects.get(exam_id=int(exam_id))
			no_of_days_present = request.POST.get('no_of_days_present')
			no_of_days_absent = request.POST.get('no_of_days_absent')
			percentage = request.POST.get('percentage')
			reg = request.POST.get('reg_no')
			class_id = Class.objects.get(class_id=cls_id)
			section_id = Section.objects.get(section_id=sec_id)
			crt=Attendance.objects.create(total_days=total_days,class_id=class_id,section_id=section_id,no_of_days_present=no_of_days_present,
			exam_id=ex,no_of_days_absent=no_of_days_absent,percentage=percentage,reg_no=reg,student_id=sid) 
			if crt:
				messages.success(request,"Added Successfully")
		return render(request, 'class.html', {})
	else:
		return redirect('adminlogin')
def view_subject_exmark(request,cls_id,sec_id,pk):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		exam_id = request.GET.get('exam_id')
		cursor = connection.cursor()
		Mark = ''' SELECT s.subject_name,m.mark,m.mark_type from app_mark as m INNER JOIN app_subject as s ON m.subject_id=s.subject_id where
		m.class_id='%d' AND m.section_id='%d' AND m.student_id_id='%d' AND m.exam_id_id='%d' AND m.mark_type='%s' ''' % (cls_id,sec_id,pk,int(exam_id),'External')
		Mark_id = cursor.execute(Mark)
		marks = cursor.fetchall()
		post = ''' SELECT AVG(m.mark) from app_mark as m INNER JOIN app_subject as s ON m.subject_id=s.subject_id where
		m.class_id='%d' AND m.section_id='%d' AND m.student_id_id='%d' AND m.exam_id_id='%d' AND m.mark_type='%s' ''' % (cls_id,sec_id,pk,int(exam_id),'External')
		sql= cursor.execute(post)
		rows = cursor.fetchall()
		exam_name = Exam.objects.filter(exam_id=int(exam_id))
		return render(request, 'view_subject_exmark.html', {'marks':marks,'exam_name':exam_name,'row':rows})
	else:
		return redirect('adminlogin')
def view_attendance(request,pk):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		exam_id = request.GET.get('exam_id')
		detail = Attendance.objects.filter(exam_id=int(exam_id),student_id=pk)
		exam_name = Exam.objects.filter(exam_id=int(exam_id))
		return render(request, 'view_attendance.html', {'detail':detail,'exam_name':exam_name})
	else:
		return redirect('adminlogin')
def view_average(request,cls_id,sec_id,pk):
	if request.session.has_key('username') or request.session.has_key('teacher') or request.session.has_key('student'):
		exam_id = request.GET.get('exam_id')
		cursor = connection.cursor()
		Mark = ''' SELECT AVG(m.mark) from app_mark as m INNER JOIN app_subject as s ON m.subject_id=s.subject_id where
		m.class_id='%d' AND m.section_id='%d' AND m.student_id_id='%d' AND m.exam_id_id='%d' AND m.mark_type='%s' ''' % (cls_id,sec_id,pk,int(exam_id),'Internal')
		Mark_id = cursor.execute(Mark)
		marks_avg = cursor.fetchall()
		post = ''' SELECT AVG(m.mark) from app_mark as m INNER JOIN app_subject as s ON m.subject_id=s.subject_id where
		m.class_id='%d' AND m.section_id='%d' AND m.student_id_id='%d' AND m.exam_id_id='%d' AND m.mark_type='%s' ''' % (cls_id,sec_id,pk,int(exam_id),'External')
		sql= cursor.execute(post)
		avg_mark = cursor.fetchall()
		attendance = Attendance.objects.filter(student_id=pk,exam_id=int(exam_id))
		exam_name = Exam.objects.filter(exam_id=int(exam_id))
		return render(request, 'view_average.html', {'marks':marks_avg,'exam_name':exam_name,'row':avg_mark,'attendance':attendance})
	else:
		return redirect('adminlogin')

def single(request):
	if request.session.has_key('student'):
		reg_no = request.session['register_number']
		cursor = connection.cursor()
		Mark = ''' SELECT s.name,s.reg_no,m.class_id_id,m.section_id_id,m.student_id_id from app_studentlogin as s INNER JOIN app_mark as d ON s.reg_no=d.reg_no INNER JOIN 
		app_studentsection as m ON d.student_id_id=m.student_id_id where s.reg_no='%s' GROUP BY s.reg_no''' % (reg_no)
		Mark_id = cursor.execute(Mark)
		detail = cursor.fetchall()
		return render(request, 'single.html', {'detail':detail})
	else:
		return redirect('adminlogin')
def sexam(request,cls_id,sec_id,pk):
	if request.session.has_key('student'):
		detail = Exam.objects.filter(class_id=cls_id,section_id=sec_id)
		return render(request, 'sexam.html', {'detail':detail,'pk':pk})
	else:
		return redirect('adminlogin')

def view_all(request,cls_id,sec_id,pk,stud_id):
	if request.session.has_key('student'):
		return render(request, 'view_all.html', {'cls_id':cls_id,'sec_id':sec_id,'stud_id':stud_id,'pk':pk})
	else:
		return redirect('adminlogin')
def register(request):
	if request.method == 'POST':
		name = request.POST.get('name')
		reg_no = request.POST.get('reg_no')
		email = request.POST.get('email')
		mobile = request.POST.get('mobile')
		city = request.POST.get('city')
		state = request.POST.get('state')
		country = request.POST.get('country')
		username = request.POST.get('username')
		password = request.POST.get('password')
		address = request.POST.get('address')
		crt=StudentLogin.objects.create(name=name,reg_no=reg_no,email=email,mobile=mobile,city=city,state=state,country=country,
		username=username,password=password,address=address)
		if crt:
			messages.success(request, 'Registered Successfully')
	return render(request, 'register.html', {})
def profile(request):
	if request.session.has_key('student'):
		ids = request.session['register_number']
		detail = StudentDetail.objects.filter(register_number=ids)
		return render(request, 'profile.html', {'detail':detail})
	else:
		return redirect('student_login')
def exportcsv(request):
	academic_year = request.session['academic_year']
	cls_id = request.GET.get('cls_id')
	sec_id = request.GET.get('sec_id')
	students = StudentSection.objects.filter(class_id=int(cls_id),section_id=int(sec_id),academic_year=academic_year)
	response = HttpResponse('')
	response['Content-Disposition'] = 'attachment; filename=students.csv'
	writer = csv.writer(response)
	writer.writerow(['Register Number', 'Student Name'])
	studs = students.values_list('reg_no','student_name')
	for std in studs:
		writer.writerow(std)
	return response